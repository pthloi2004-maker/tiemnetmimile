import csv
import io
import json
import os
import random
import string
import traceback

from django.contrib import messages
from django.contrib.auth import login, logout, authenticate, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.db.models import Sum, Count, Q
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.core.files.uploadedfile import UploadedFile
from django.conf import settings
from django.core.validators import validate_email
from django.core.exceptions import ValidationError

from .forms import (
    CheckinForm, CustomerForm, QueueForm, ServiceOrderForm,
    SchedulingForm, MachineForm, PhoneRequestForm, OTPVerifyForm,
)
from .models import Machine, Session, SessionService, Service, Customer, Queue, OTPCode, LoginToken
from .os_simulators import (
    build_memory_reference_stream,
    build_memory_comparison,
    parse_reference_string,
    simulate_bounded_buffer,
    simulate_cpu_memory,
)
from .scheduling_algorithms import (
    sessions_to_processes, queue_to_processes,
    fcfs, sjf, round_robin,
    build_banker_data, banker_safety, banker_request,
    calc_stats, RESOURCE_NAMES, TOTAL_RESOURCES,
    renumber_processes, generate_random_processes, normalize_arrival_times,
    build_scheduling_events, ALGORITHM_DESCRIPTIONS,
    dbdaa_safety, dbdaa_request,
)
from .omdrrs_scheduler import schedule_omdrrs
from .scheduling_benchmark import run_all_schedulers
from .utils import role_required


def get_rate_for(moment):
    hour = moment.hour
    if 6 <= hour < 18:
        return 8000
    if hour >= 22 or hour < 6:
        return 6000
    return 7000


def expire_overdue_sessions():
    """Đóng các phiên đang chạy nếu đã vượt quá thời lượng dự kiến."""
    active_sessions = Session.objects.filter(status='dang_chay')
    for sess in active_sessions:
        sess.expire_if_overdue()


# ==============================================================
# TRANG CHỦ — DASHBOARD
# ==============================================================
@login_required(login_url='app:login')
def dashboard(request):
    expire_overdue_sessions()
    machines       = Machine.objects.all()
    active_sessions_qs = Session.objects.filter(status='dang_chay').select_related('machine', 'customer')
    waiting_queue  = Queue.objects.filter(is_served=False).order_by('arrived_at')

    # Thống kê hôm nay
    today = timezone.now().date()
    today_sessions = Session.objects.filter(start_time__date=today)
    paid_revenue = today_sessions.filter(paid=True).aggregate(
        total=Sum('total_cost')
    )['total'] or 0
    active_revenue = sum(
        (sess.calculate_cost() + sess.get_service_cost()) for sess in active_sessions_qs
        if sess.start_time.date() == today
    )
    revenue_today  = paid_revenue + active_revenue

    # Tổng doanh thu từ trước tới giờ: các phiên đã hoàn thành + phiên đang chạy hiện tại
    historical_revenue = Session.objects.exclude(status='dang_chay').aggregate(total=Sum('total_cost'))['total'] or 0
    active_revenue_all = sum(
        (sess.calculate_cost() + sess.get_service_cost()) for sess in active_sessions_qs
    )
    total_revenue = historical_revenue + active_revenue_all

    context = {
        'machines':         machines,
        'total_machines':   machines.count(),
        'customer_count':   Customer.objects.count(),
        'active_sessions':  active_sessions_qs.count(),
        'active_revenue':   active_revenue,
        'waiting_queue':    waiting_queue,
        'revenue_today':    revenue_today,
        'total_revenue':    total_revenue,
        'today':            today,
        'today_count':      today_sessions.count(),
        'today_sessions':   today_sessions.order_by('-start_time')[:6],
        'active_sessions_details': active_sessions_qs.order_by('start_time')[:8],
        'machines_in_use':  active_sessions_qs.values('machine').distinct().count(),
        'machines_avail':   machines.filter(status='trong').count(),
        'machines_maint':  machines.filter(status='bao_tri').count(),
        'machines_error':   machines.filter(status='loi').count(),
        'machines_locked':  machines.filter(status='khoa').count(),
        'queue_count':      waiting_queue.count(),
    }
    return render(request, 'app/dashboard.html', context)



# ==============================================================
# MÁY TÍNH
# ==============================================================
@login_required(login_url='app:login')
def machine_list(request):
    machines = Machine.objects.all()
    return render(request, 'app/machine/list.html', {'machines': machines})


@login_required(login_url='app:login')
def machine_detail(request, machine_id):
    machine = get_object_or_404(Machine, id=machine_id)
    sessions = Session.objects.filter(machine=machine).order_by('-start_time')[:20]
    return render(request, 'app/machine/detail.html', {
        'machine': machine,
        'sessions': sessions,
    })


@role_required(['owner', 'staff', 'tech'])
def machine_add(request):
    if request.method == 'POST':
        form = MachineForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã thêm máy mới thành công.')
            return redirect('app:machine_list')
    else:
        form = MachineForm()
    return render(request, 'app/machine/form.html', {
        'form': form,
        'title': 'Thêm Máy Mới',
        'submit_label': 'Lưu Máy',
    })


@role_required(['owner', 'staff', 'tech'])
def machine_edit(request, machine_id):
    machine = get_object_or_404(Machine, id=machine_id)
    if request.method == 'POST':
        form = MachineForm(request.POST, request.FILES, instance=machine)
        if form.is_valid():
            form.save()
            messages.success(request, f'Đã cập nhật {machine.name}.')
            return redirect('app:machine_detail', machine_id=machine.id)
    else:
        form = MachineForm(instance=machine)
    return render(request, 'app/machine/form.html', {
        'form': form,
        'machine': machine,
        'title': f'Chỉnh Sửa Máy {machine.name}',
        'submit_label': 'Lưu Thay Đổi',
    })


@require_POST
@role_required(['owner', 'staff', 'tech'])
def machine_delete(request, machine_id):
    machine = get_object_or_404(Machine, id=machine_id)
    if Session.objects.filter(machine=machine).exists():
        messages.error(
            request,
            f'Không thể xóa {machine.name} vì máy đã có lịch sử phiên. Hãy khóa máy thay vì xóa.'
        )
        return redirect('app:machine_detail', machine_id=machine.id)

    name = machine.name
    machine.delete()
    messages.success(request, f'Đã xóa máy {name}.')
    return redirect('app:machine_list')


@role_required(['owner', 'staff', 'tech'])
def machine_toggle_lock(request, machine_id):
    machine = get_object_or_404(Machine, id=machine_id)
    if machine.is_locked():
        machine.status = 'trong'
        action = 'mở khóa'
    else:
        machine.status = 'khoa'
        action = 'khóa'
    machine.save()
    messages.success(request, f'✅ Đã {action} {machine.name}.')
    return redirect('app:machine_detail', machine_id=machine.id)


@login_required
def api_machines(request):
    machines = Machine.objects.all().order_by('name')
    data = []
    for m in machines:
        active = Session.objects.filter(machine=m, status='dang_chay').order_by('-start_time').first()
        item = {
            'id': m.id,
            'name': m.name,
            'type': m.machine_type,
            'status': m.status,
            'hourly_rate': float(m.hourly_rate),
            'has_headset': m.has_headset,
            'has_account': m.has_account,
            'ram_gb': m.ram_gb,
            'image_url': m.image.url if m.image else None,
            'active': None,
        }
        if active:
            item['active'] = {
                'session_id': active.id,
                'customer': active.get_customer_name(),
                'game': active.game_name,
                'started_at': active.start_time.isoformat(),
                'remaining_minutes': active.get_remaining_minutes(),
            }
            item['status'] = 'dang_dung'
        elif m.status == 'dang_dung':
            # Correct stale machine status in the API output if there is no active session.
            item['status'] = 'trong'
        data.append(item)
    summary = {
        'total': machines.count(),
        'in_use': Session.objects.filter(status='dang_chay').values('machine').distinct().count(),
        'available': machines.filter(status='trong').count(),
        'maintenance': machines.filter(status='bao_tri').count(),
        'error': machines.filter(status='loi').count(),
        'locked': machines.filter(status='khoa').count(),
    }
    return JsonResponse({'machines': data, 'summary': summary})


@require_POST
@role_required(['owner', 'staff', 'tech'])
def api_machine_action(request, machine_id):
    action = request.POST.get('action')
    machine = get_object_or_404(Machine, id=machine_id)
    if action == 'lock':
        machine.status = 'khoa'
        machine.save()
        return JsonResponse({'ok': True, 'status': 'khoa'})
    if action == 'unlock':
        machine.status = 'trong'
        machine.save()
        return JsonResponse({'ok': True, 'status': 'trong'})
    if action == 'checkout':
        # Force checkout current session if exists
        sess = Session.objects.filter(machine=machine, status='dang_chay').order_by('-start_time').first()
        if sess:
            sess.end_time = timezone.now()
            sess.total_cost = sess.calculate_cost()
            sess.status = 'hoan_thanh'
            sess.paid = False
            sess.save()
        machine.status = 'trong'
        machine.save()
        return JsonResponse({'ok': True})
    return JsonResponse({'ok': False, 'error': 'unknown_action'}, status=400)


# ==============================================================
# CHECK-IN
# ==============================================================
@login_required(login_url='app:login')
def checkin(request, machine_id=None):
    """
    Nhận khách vào máy. Trước khi cấp phát, kiểm tra Banker Algorithm.
    Nếu không còn máy trống, đề xuất thêm vào hàng chờ.
    """
    expire_overdue_sessions()
    initial = {}
    if machine_id:
        machine = get_object_or_404(Machine, id=machine_id, status='trong')
        initial['machine'] = machine

    # Kiểm tra xem còn máy trống không
    available_machines = Machine.objects.filter(status='trong')
    no_machines_available = not available_machines.exists()

    if request.method == 'POST':
        # Nếu là thêm vào hàng chờ từ form checkin
        if 'add_to_queue' in request.POST:
            customer_name = request.POST.get('customer_name', '').strip()
            phone = request.POST.get('phone', '').strip()
            preferred_type = request.POST.get('preferred_type', 'thuong')
            planned_minutes = int(request.POST.get('planned_minutes', 60))
            note = request.POST.get('note', '').strip()
            if customer_name:
                Queue.objects.create(
                    customer_name=customer_name,
                    phone=phone,
                    preferred_type=preferred_type,
                    planned_minutes=planned_minutes,
                    note=note,
                )
                messages.success(request, f'✅ Đã thêm {customer_name} vào hàng chờ! Khi có máy trống sẽ phục vụ ngay.')
            else:
                messages.error(request, '❌ Vui lòng nhập tên khách.')
            return redirect('app:dashboard')

        form = CheckinForm(request.POST)
        if form.is_valid():
            sess = form.save(commit=False)
            sess.user = request.user
            sess.status = 'dang_chay'

            # ---- Kiểm tra Banker trước khi cấp phát ----
            active_sessions = list(Session.objects.filter(status='dang_chay'))
            banker_data     = build_banker_data(active_sessions)

            request_vec = [
                1 if sess.used_headset else 0,
                1 if sess.used_account else 0,
                sess.used_ram_gb,
            ]
            avail = banker_data['available']
            enough = all(request_vec[j] <= avail[j] for j in range(len(avail)))

            if not enough:
                messages.error(request,
                    '❌ Không đủ tài nguyên khả dụng! '
                    f'Tai nghe còn: {avail[0]}, Tài khoản còn: {avail[1]}, RAM còn: {avail[2]} GB.'
                )
                return render(request, 'app/session/checkin.html', {
                    'form': form,
                    'machine_options': Machine.objects.filter(status='trong'),
                    'customer_options': Customer.objects.all(),
                    'services': Service.objects.filter(available=True).order_by('category', 'name'),
                    'current_rate': get_rate_for(timezone.now()),
                    'no_machines_available': no_machines_available,
                    'queue_count': Queue.objects.filter(is_served=False).count(),
                })

            # Lưu phiên
            sess.save()

            # Lưu dịch vụ chọn thêm
            service_ids = request.POST.getlist('services')
            added_services = 0
            for service in Service.objects.filter(id__in=service_ids, available=True):
                qty = int(request.POST.get(f'quantity_{service.id}', '0') or 0)
                if qty > 0:
                    SessionService.objects.create(
                        session=sess,
                        service=service,
                        quantity=qty,
                        price=service.price,
                    )
                    added_services += qty

            # Cập nhật trạng thái máy
            machine = sess.machine
            machine.status = 'dang_dung'
            machine.save()

            success_msg = f'✅ Đã nhận {sess.get_customer_name()} vào {machine.name}!'
            if added_services:
                success_msg += f' (Đã thêm {added_services} món/dịch vụ).'
            messages.success(request, success_msg)
            return redirect('app:dashboard')
    else:
        form = CheckinForm(initial=initial)

    # Hiển thị tình trạng tài nguyên hiện tại (cho nhân viên biết)
    active_sessions = list(Session.objects.filter(status='dang_chay'))
    banker_data     = build_banker_data(active_sessions)

    # Tính thời gian chờ dự kiến cho khách mới
    queue_count = Queue.objects.filter(is_served=False).count()
    estimated_wait = 0
    if no_machines_available and active_sessions:
        total_remaining = sum(s.get_remaining_minutes() for s in active_sessions)
        total_machines = Machine.objects.filter(status__in=['trong', 'dang_dung']).count()
        busy = Machine.objects.filter(status='dang_dung').count()
        free = max(1, total_machines - busy)
        estimated_wait = (total_remaining + queue_count * 60) // free

    return render(request, 'app/session/checkin.html', {
        'form':             form,
        'banker_data':      banker_data,
        'resource_names':   RESOURCE_NAMES,
        'machine_options':  available_machines,
        'customer_options': Customer.objects.all(),
        'services':         Service.objects.filter(available=True).order_by('category', 'name'),
        'current_rate':     get_rate_for(timezone.now()),
        'no_machines_available': no_machines_available,
        'queue_count':      queue_count,
        'estimated_wait':   estimated_wait,
    })

def _pick_queue_for_machine(machine):
    next_queue = Queue.objects.filter(is_served=False).order_by('arrived_at').first()
    if not next_queue:
        return None

    preferred = Queue.objects.filter(
        is_served=False,
        preferred_machine=machine,
    ).order_by('arrived_at').first()
    if preferred:
        return preferred

    type_match = Queue.objects.filter(
        is_served=False,
        preferred_type=machine.machine_type,
    ).order_by('arrived_at').first()
    if type_match:
        return type_match

    return next_queue


def _start_session_from_queue(queue_item, machine, user):
    session = Session.objects.create(
        user=user,
        customer_name=queue_item.customer_name,
        machine=machine,
        planned_minutes=queue_item.planned_minutes,
        status='dang_chay',
    )
    queue_item.is_served = True
    queue_item.save()
    machine.status = 'dang_dung'
    machine.save()
    return session


# ==============================================================
# CHECK-OUT — TRẢ MÁY
# ==============================================================
@login_required(login_url='app:login')
def checkout(request, session_id):
    sess = get_object_or_404(Session, id=session_id, status='dang_chay')

    sess.end_time   = timezone.now()
    sess.total_cost = sess.calculate_cost()
    service_cost = sum(s.get_cost() for s in sess.services.all())
    sess.total_cost += service_cost
    sess.status     = 'hoan_thanh'
    sess.paid       = True
    sess.save()

    machine = sess.machine
    machine.status = 'trong'
    machine.save()

    next_queue = _pick_queue_for_machine(machine)
    next_session = None
    if next_queue:
        next_session = _start_session_from_queue(next_queue, machine, request.user)

    messages.success(request,
        f'? {sess.get_customer_name()} ?? tr? m?y {machine.name}. '
        f'T?ng ti?n: {sess.total_cost:,}?'
    )

    if next_session:
        messages.success(request,
            f'?? ?? t? ??ng ??a kh?ch {next_session.get_customer_name()} t? h?ng ch? v?o {machine.name}.'
        )
        return redirect('app:session_detail', session_id=next_session.id)

    return redirect('app:session_detail', session_id=sess.id)





# ==============================================================
# DANH SÁCH & CHI TIẾT PHIÊN
# ==============================================================
@login_required(login_url='app:login')
def session_list(request):
    status  = request.GET.get('status', '')
    sessions = Session.objects.select_related('machine', 'customer').order_by('-start_time')
    if status:
        sessions = sessions.filter(status=status)
    return render(request, 'app/session/list.html', {
        'sessions':        sessions,
        'selected_status': status,
    })


@login_required(login_url='app:login')
def session_detail(request, session_id):
    sess = get_object_or_404(Session, id=session_id)
    services = sess.services.select_related('service').all()
    service_total = sum(item.get_cost() for item in services)
    play_cost = sess.calculate_cost() if sess.status == 'dang_chay' else sess.total_cost
    total_cost = round(play_cost + service_total)
    return render(request, 'app/session/detail.html', {
        'session': sess,
        'session_services': services,
        'service_total': service_total,
        'play_cost': play_cost,
        'total_cost': total_cost,
    })


# ==============================================================
# DỊCH VỤ
# ==============================================================
@login_required(login_url='app:login')
def service_list(request):
    services = Service.objects.all()
    return render(request, 'app/service/list.html', {'services': services})


@login_required(login_url='app:login')
def order_service(request, session_id):
    sess = get_object_or_404(Session, id=session_id, status='dang_chay')
    if request.method == 'POST':
        form = ServiceOrderForm(request.POST)
        if form.is_valid():
            svc = form.cleaned_data['service']
            qty = form.cleaned_data['quantity']
            SessionService.objects.create(
                session=sess, service=svc,
                quantity=qty, price=svc.price
            )
            messages.success(request, f'Đã thêm {svc.name} × {qty} vào phiên.')
            return redirect('app:session_detail', session_id=sess.id)
    else:
        form = ServiceOrderForm()

    services = Service.objects.filter(available=True).order_by('category', 'name')
    category_labels = dict(Service.CATEGORY_CHOICES)
    return render(request, 'app/service/order.html', {
        'form': form,
        'session': sess,
        'services': services,
        'category_labels': category_labels,
    })


# ==============================================================
# HÀNG CHỜ
# ==============================================================
@login_required(login_url='app:login')
def queue_view(request):
    waiting = Queue.objects.filter(is_served=False).order_by('arrived_at')
    
    # Tính thời gian chờ dự kiến
    active_sessions = Session.objects.filter(status='dang_chay')
    total_remaining = sum(s.get_remaining_minutes() for s in active_sessions)
    total_machines = Machine.objects.filter(status__in=['trong', 'dang_dung']).count()
    busy = Machine.objects.filter(status='dang_dung').count()
    free = max(1, total_machines - busy)
    estimated_wait = (total_remaining + waiting.count() * 60) // free if waiting.exists() else 0
    
    machines_available = Machine.objects.filter(status='trong').count()
    
    return render(request, 'app/queue/list.html', {
        'queue': waiting,
        'machines_available': machines_available,
        'estimated_wait': estimated_wait,
    })


@login_required(login_url='app:login')
def queue_add(request):
    if request.method == 'POST':
        form = QueueForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã thêm khách vào hàng chờ.')
            return redirect('app:queue')
    else:
        form = QueueForm()
    return render(request, 'app/queue/add.html', {'form': form})


@login_required(login_url='app:login')
def queue_serve(request, queue_id):
    q = get_object_or_404(Queue, id=queue_id, is_served=False)

    available = Machine.objects.filter(status='trong')
    preferred_machine = None

    if q.preferred_machine:
        if q.preferred_machine.status == 'trong':
            preferred_machine = q.preferred_machine
        else:
            messages.warning(request,
                f'?? {q.customer_name} y?u c?u {q.preferred_machine.name}, nh?ng m?y n?y ch?a tr?ng. Kh?ch v?n n?m trong h?ng ch?.'
            )
            return redirect('app:queue')
    elif q.preferred_type:
        preferred_machine = available.filter(machine_type=q.preferred_type).first()
    else:
        preferred_machine = available.first()

    if not preferred_machine:
        messages.warning(request,
            f'?? Hi?n kh?ng c? m?y tr?ng ph? h?p cho {q.customer_name}. Kh?ch v?n n?m trong h?ng ch?.'
        )
        return redirect('app:queue')

    session = _start_session_from_queue(q, preferred_machine, request.user)
    messages.success(request,
        f'? ?? ??a kh?ch {session.get_customer_name()} v?o {preferred_machine.name} t? h?ng ch?.'
    )
    return redirect('app:session_detail', session_id=session.id)


# ==============================================================
# KHÁCH HÀNG
# ==============================================================
@login_required(login_url='app:login')
def customer_list(request):
    query = request.GET.get('q', '')
    customers = Customer.objects.all()
    if query:
        customers = customers.filter(
            Q(name__icontains=query) | Q(phone__icontains=query)
        )
    return render(request, 'app/customer/list.html', {
        'customers': customers,
        'query':     query,
    })


@login_required(login_url='app:login')
def customer_add(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã thêm khách hàng thành viên.')
            return redirect('app:customer_list')
    else:
        form = CustomerForm()
    return render(request, 'app/customer/add.html', {'form': form})


@login_required(login_url='app:login')
def customer_detail(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    sessions = Session.objects.filter(customer=customer).order_by('-start_time')
    return render(request, 'app/customer/detail.html', {
        'customer': customer,
        'sessions': sessions,
    })


# ==============================================================
# MÔ PHỎNG HỆ ĐIỀU HÀNH — TRANG TỔNG HỢP
# ==============================================================
@login_required(login_url='app:login')
def simulation(request):
    """Trang giới thiệu tổng quan mô phỏng"""
    expire_overdue_sessions()
    active_sessions = Session.objects.filter(status='dang_chay').select_related('machine', 'customer').order_by('start_time')[:8]
    waiting_queue = Queue.objects.filter(is_served=False).order_by('arrived_at')[:8]
    machine_status = {
        'total': Machine.objects.count(),
        'in_use': Session.objects.filter(status='dang_chay').values('machine').distinct().count(),
        'available': Machine.objects.filter(status='trong').count(),
        'maintenance': Machine.objects.filter(status='bao_tri').count(),
    }
    return render(request, 'app/simulation/index.html', {
        'active_count': active_sessions.count(),
        'queue_count':  waiting_queue.count(),
        'active_sessions': active_sessions,
        'waiting_queue': waiting_queue,
        'machine_status': machine_status,
    })


@login_required(login_url='app:login')
def import_simulation_data(request):
    """
    Import simulation data from JSON or CSV file.
    Supports: scheduling, banker, memory, synchronization
    GET: Render import page with drag & drop, preview, validation
    POST: Process uploaded file
    """
    import json
    import csv
    import io
    
    # --- GET: Render import page ---
    if request.method == 'GET':
        return render(request, 'app/simulation/import_simulation.html', {
            'title': 'Nhập Dữ Liệu Mô Phỏng',
        })
    
    # --- POST: Process uploaded file ---
    if 'file' not in request.FILES:
        messages.error(request, '❌ Vui lòng chọn một file dữ liệu (JSON hoặc CSV).')
        return redirect('app:simulation')
    
    file = request.FILES['file']
    
    # --- File size validation (max 10MB) ---
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    if file.size > MAX_FILE_SIZE:
        messages.error(request, f'❌ Dung lượng file quá lớn ({file.size / 1024 / 1024:.1f}MB). Tối đa 10MB.')
        return redirect('app:simulation')
    
    # --- Empty file check ---
    if file.size == 0:
        messages.error(request, '❌ File rỗng. Vui lòng chọn file có dữ liệu.')
        return redirect('app:simulation')
    
    # --- Detect file type by extension ---
    filename = file.name.lower()
    is_csv = filename.endswith('.csv')
    is_json = filename.endswith('.json')
    
    if not is_csv and not is_json:
        messages.error(request, '❌ Định dạng file không hỗ trợ. Chỉ chấp nhận file .json hoặc .csv.')
        return redirect('app:simulation')
    
    # --- Parse file ---
    data = None
    try:
        if is_json:
            data = json.load(file)
        elif is_csv:
            # Read CSV content with BOM handling
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
            if not rows:
                messages.error(request, '❌ File CSV rỗng hoặc không có dữ liệu.')
                return redirect('app:simulation')
            
            # Detect type from CSV columns
            headers = reader.fieldnames or []
            headers_lower = [h.lower().strip() for h in headers]
            
            # Check for scheduling columns
            if all(k in headers_lower for k in ['arrival', 'burst']):
                processes = []
                for i, row in enumerate(rows):
                    try:
                        processes.append({
                            'pid': int(row.get('pid', i + 1)),
                            'name': str(row.get('name', f"P{row.get('pid', i + 1)}")),
                            'arrival': int(row.get('arrival', 0)),
                            'burst': int(row.get('burst', 10)),
                            'priority': int(row.get('priority', 0)),
                        })
                    except (ValueError, TypeError) as e:
                        messages.error(request, f'❌ Lỗi dữ liệu tại dòng {i+2}: {str(e)}')
                        return redirect('app:simulation')
                data = {
                    'type': 'scheduling',
                    'processes': processes,
                    'quantum': 20
                }
            elif all(k in headers_lower for k in ['allocation', 'max_need']):
                total = []
                allocation = []
                max_need = []
                names = []
                for row in rows:
                    try:
                        alloc_str = str(row.get('allocation', '')).strip()
                        max_str = str(row.get('max_need', '')).strip()
                        alloc_vals = [int(x.strip()) for x in alloc_str.replace('[','').replace(']','').split(',') if x.strip()]
                        max_vals = [int(x.strip()) for x in max_str.replace('[','').replace(']','').split(',') if x.strip()]
                        if not total and 'total' in headers_lower:
                            total_str = str(row.get('total', '')).strip()
                            total = [int(x.strip()) for x in total_str.replace('[','').replace(']','').split(',') if x.strip()]
                        allocation.append(alloc_vals)
                        max_need.append(max_vals)
                        names.append(str(row.get('name', f"P{len(names)+1}")))
                    except (ValueError, TypeError) as e:
                        messages.error(request, f'❌ Lỗi dữ liệu Banker: {str(e)}')
                        return redirect('app:simulation')
                data = {
                    'type': 'banker',
                    'banker': {
                        'total': total or [10, 5, 12],
                        'allocation': allocation,
                        'max_need': max_need,
                        'names': names
                    }
                }
            elif 'references' in headers_lower:
                refs = []
                for row in rows:
                    ref_str = str(row.get('references', '')).strip()
                    refs.extend([int(x.strip()) for x in ref_str.split(',') if x.strip()])
                frame_count = int(rows[0].get('frame_count', 3)) if rows else 3
                data = {
                    'type': 'memory',
                    'memory': {
                        'references': refs,
                        'frame_count': max(2, frame_count)
                    }
                }
            elif 'sequence' in headers_lower:
                seq = []
                for row in rows:
                    s = str(row.get('sequence', '')).strip().upper()
                    if s in ('P', 'C', 'PRODUCER', 'CONSUMER'):
                        seq.append(s[0])
                buffer_size = int(rows[0].get('buffer_size', 3)) if rows else 3
                data = {
                    'type': 'synchronization',
                    'synchronization': {
                        'sequence': seq,
                        'buffer_size': max(1, buffer_size),
                        'producer_count': 3,
                        'consumer_count': 2
                    }
                }
            else:
                messages.error(request, '❌ Không thể xác định loại dữ liệu từ file CSV. Các cột cần có: arrival, burst (scheduling) hoặc allocation, max_need (banker) hoặc references (memory) hoặc sequence (sync).')
                return redirect('app:simulation')
    except json.JSONDecodeError as e:
        messages.error(request, f'❌ Lỗi đọc file JSON: {str(e)}. Vui lòng kiểm tra định dạng JSON.')
        return redirect('app:simulation')
    except csv.Error as e:
        messages.error(request, f'❌ Lỗi đọc file CSV: {str(e)}.')
        return redirect('app:simulation')
    except Exception as e:
        messages.error(request, f'❌ Lỗi xử lý file: {str(e)}')
        return redirect('app:simulation')
    
    if data is None:
        messages.error(request, '❌ Không thể đọc dữ liệu từ file.')
        return redirect('app:simulation')
    
    sim_type = data.get('type')
    if not sim_type:
        # Auto-detect type from keys
        if 'processes' in data:
            sim_type = 'scheduling'
        elif 'banker' in data:
            sim_type = 'banker'
        elif 'memory' in data:
            sim_type = 'memory'
        elif 'synchronization' in data:
            sim_type = 'synchronization'
        else:
            messages.error(request, '❌ Không thể xác định loại dữ liệu mô phỏng. File cần có trường "type" hoặc chứa processes/banker/memory/synchronization.')
            return redirect('app:simulation')
            
    redirect_to_target = request.POST.get('redirect_to', '')
            
    if sim_type == 'scheduling':
        processes = data.get('processes', [])
        if not processes:
            messages.error(request, '❌ Dữ liệu scheduling thiếu danh sách processes.')
            return redirect('app:simulation')
            
        try:
            valid_processes = []
            for idx, p in enumerate(processes):
                valid_processes.append({
                    'pid': int(p.get('pid', idx + 1)),
                    'name': str(p.get('name', f"P{idx+1}")),
                    'arrival': max(0, int(p.get('arrival', 0))),
                    'burst': max(1, int(p.get('burst', 5))),
                    'priority': int(p.get('priority', 3))
                })
        except (ValueError, TypeError) as e:
            messages.error(request, f'❌ Lỗi kiểu dữ liệu trong processes: {str(e)}')
            return redirect('app:simulation')
            
        if not valid_processes:
            messages.error(request, '❌ Không có tiến trình hợp lệ trong file.')
            return redirect('app:simulation')
            
        quantum = data.get('quantum', 20)
        try:
            quantum = max(1, int(quantum))
        except (ValueError, TypeError):
            quantum = 20
            
        request.session['custom_processes'] = valid_processes
        request.session['custom_quantum'] = quantum
        messages.success(request, f'✅ Đã nhập {len(valid_processes)} tiến trình từ file.')

        # Redirect to appropriate simulation page and ensure it loads file mode
        if redirect_to_target == 'app:scheduling_compare':
            return redirect(f"{reverse('app:scheduling_compare')}?data_mode=file")
        elif redirect_to_target == 'app:omdrrs_simulation':
            return redirect(f"{reverse('app:omdrrs_simulation')}?data_mode=file")
        return redirect(f"{reverse('app:scheduling_simulation')}?data_mode=file")
        
    elif sim_type == 'banker':
        banker_data = data.get('banker', {})
        total = banker_data.get('total')
        allocation = banker_data.get('allocation')
        max_need = banker_data.get('max_need')
        names = banker_data.get('names')
        
        if not total or not allocation or not max_need:
            messages.error(request, '❌ Dữ liệu Banker thiếu total, allocation hoặc max_need.')
            return redirect('app:simulation')
            
        try:
            total = [int(x) for x in total]
            allocation = [[int(x) for x in row] for row in allocation]
            max_need = [[int(x) for x in row] for row in max_need]
            if names:
                names = [str(x) for x in names]
            else:
                names = [f"P{i}" for i in range(len(allocation))]
        except (ValueError, TypeError) as e:
            messages.error(request, f'❌ Lỗi kiểu dữ liệu trong file Banker: {str(e)}')
            return redirect('app:simulation')
            
        request.session['custom_banker'] = {
            'total': total,
            'allocation': allocation,
            'max_need': max_need,
            'names': names
        }
        messages.success(request, '✅ Đã nhập thành công dữ liệu Banker Algorithm.')
        # Redirect to banker/dbdaa pages in file mode
        if redirect_to_target == 'app:dbdaa_simulation':
            return redirect(f"{reverse('app:dbdaa_simulation')}?scenario=file")
        elif redirect_to_target == 'app:banker_dbdaa_compare':
            return redirect(f"{reverse('app:banker_dbdaa_compare')}?scenario=file")
        return redirect(f"{reverse('app:banker_simulation')}?scenario=file")
        
    elif sim_type == 'memory':
        mem_data = data.get('memory', {})
        references = mem_data.get('references')
        frame_count = mem_data.get('frame_count', 3)
        if not references:
            messages.error(request, '❌ Dữ liệu bộ nhớ thiếu references.')
            return redirect('app:simulation')
            
        try:
            references = [int(x) for x in references]
            frame_count = max(2, int(frame_count))
        except (ValueError, TypeError) as e:
            messages.error(request, f'❌ Lỗi kiểu dữ liệu trong file bộ nhớ: {str(e)}')
            return redirect('app:simulation')
            
        request.session['custom_memory'] = {
            'references': references,
            'frame_count': frame_count
        }
        messages.success(request, '✅ Đã nhập thành công dữ liệu chuỗi trang bộ nhớ.')
        return redirect(f"{reverse('app:memory_simulation')}?data_mode=file")
        
    elif sim_type == 'synchronization':
        sync_data = data.get('synchronization', {})
        sequence = sync_data.get('sequence')
        buffer_size = sync_data.get('buffer_size', 3)
        producer_count = sync_data.get('producer_count', 2)
        consumer_count = sync_data.get('consumer_count', 2)
        if not sequence:
            messages.error(request, '❌ Dữ liệu đồng bộ hóa thiếu sequence.')
            return redirect('app:simulation')
            
        try:
            sequence = [str(x).strip().upper()[0] for x in sequence if str(x).strip()]
            sequence = [x for x in sequence if x in ('P', 'C')]
            buffer_size = max(1, int(buffer_size))
            producer_count = max(1, int(producer_count))
            consumer_count = max(1, int(consumer_count))
        except (ValueError, TypeError) as e:
            messages.error(request, f'❌ Lỗi kiểu dữ liệu trong file đồng bộ hóa: {str(e)}')
            return redirect('app:simulation')
            
        request.session['custom_sync'] = {
            'sequence': sequence,
            'buffer_size': buffer_size,
            'producer_count': producer_count,
            'consumer_count': consumer_count
        }
        messages.success(request, '✅ Đã nhập thành công dữ liệu đồng bộ hóa.')
        return redirect(f"{reverse('app:synchronization_simulation')}?pattern=file")
        
    elif sim_type == 'machines':
        machines = data.get('machines', [])
        if not machines:
            messages.error(request, '❌ Danh sách máy trống.')
            return redirect('app:simulation')
            
        try:
            valid_machines = []
            for idx, m in enumerate(machines):
                name = str(m.get('name', '')).strip()
                if not name:
                    continue
                machine_type = str(m.get('machine_type', 'thuong')).strip().lower()
                type_map = {
                    'thuong': 'thuong', 'máy thường': 'thuong', 'may thuong': 'thuong', 'regular': 'thuong',
                    'gaming': 'gaming', 'máy gaming': 'gaming', 'may gaming': 'gaming', 'game': 'gaming',
                    'vip': 'vip', 'phòng vip': 'vip', 'phong vip': 'vip',
                }
                machine_type = type_map.get(machine_type, 'thuong')
                
                status = str(m.get('status', 'trong')).strip().lower()
                status_map = {
                    'trong': 'trong', 'trống': 'trong', 'available': 'trong', 'free': 'trong',
                    'dang_dung': 'dang_dung', 'đang dùng': 'dang_dung', 'dang dung': 'dang_dung', 'in use': 'dang_dung', 'busy': 'dang_dung',
                    'bao_tri': 'bao_tri', 'bảo trì': 'bao_tri', 'bao tri': 'bao_tri', 'maintenance': 'bao_tri',
                    'loi': 'loi', 'lỗi': 'loi', 'error': 'loi', 'fault': 'loi',
                    'khoa': 'khoa', 'khóa': 'khoa', 'khoá': 'khoa', 'locked': 'khoa',
                }
                status = status_map.get(status, 'trong')
                
                hourly_rate = m.get('hourly_rate', '10000')
                try:
                    hourly_rate = int(float(str(hourly_rate).replace(',', '').replace('.', '')))
                    if hourly_rate < 0:
                        hourly_rate = 10000
                except (ValueError, TypeError):
                    hourly_rate = 10000
                
                has_headset = str(m.get('has_headset', 'true')).strip().lower() in ('true', 'yes', '1', 'có', 'co')
                has_account = str(m.get('has_account', 'false')).strip().lower() in ('true', 'yes', '1', 'có', 'co')
                
                ram_gb = m.get('ram_gb', '8')
                try:
                    ram_gb = max(1, int(float(ram_gb)))
                except (ValueError, TypeError):
                    ram_gb = 8
                
                note = str(m.get('note', '')).strip()
                
                valid_machines.append({
                    'name': name,
                    'machine_type': machine_type,
                    'status': status,
                    'hourly_rate': hourly_rate,
                    'has_headset': has_headset,
                    'has_account': has_account,
                    'ram_gb': ram_gb,
                    'note': note,
                })
            
            if not valid_machines:
                messages.error(request, '❌ Không có máy hợp lệ trong file.')
                return redirect('app:simulation')
                
            request.session['custom_machines'] = valid_machines
            messages.success(request, f'✅ Đã nhập {len(valid_machines)} máy từ file.')

            if redirect_to_target == 'app:omdrrs_simulation':
                return redirect(f"{reverse('app:omdrrs_simulation')}?data_mode=file")
            return redirect(f"{reverse('app:simulation')}?data_mode=file")
        except (ValueError, TypeError) as e:
            messages.error(request, f'❌ Lỗi kiểu dữ liệu trong file máy: {str(e)}')
            return redirect('app:simulation')
        
    messages.error(request, '❌ Loại mô phỏng không được hỗ trợ.')
    return redirect('app:simulation')


# ==============================================================
# Helper: Save imported data to session
# ==============================================================
def _save_imported_data(request, data, sim_type=None):
    """Save validated import data to session and return (success, messages)."""
    if not sim_type:
        sim_type = data.get('type')
    if not sim_type:
        if 'processes' in data:
            sim_type = 'scheduling'
        elif 'banker' in data:
            sim_type = 'banker'
        elif 'memory' in data:
            sim_type = 'memory'
        elif 'synchronization' in data:
            sim_type = 'synchronization'
        elif 'machines' in data:
            sim_type = 'machines'
        else:
            return False, ['Không thể xác định loại dữ liệu mô phỏng.']
    
    successes = []
    errors = []
    
    try:
        if sim_type == 'scheduling':
            processes = data.get('processes', [])
            if not processes:
                errors.append('Dữ liệu scheduling thiếu danh sách processes.')
            else:
                valid_processes = []
                for idx, p in enumerate(processes):
                    valid_processes.append({
                        'pid': int(p.get('pid', idx + 1)),
                        'name': str(p.get('name', f"P{idx+1}")),
                        'arrival': max(0, int(p.get('arrival', 0))),
                        'burst': max(1, int(p.get('burst', 5))),
                        'priority': int(p.get('priority', 3))
                    })
                quantum = data.get('quantum', 20)
                try:
                    quantum = max(1, int(quantum))
                except (ValueError, TypeError):
                    quantum = 20
                request.session['custom_processes'] = valid_processes
                request.session['custom_quantum'] = quantum
                successes.append(f'Đã nhập {len(valid_processes)} tiến trình.')
                
        elif sim_type == 'banker':
            banker_data = data.get('banker', {})
            total = banker_data.get('total')
            allocation = banker_data.get('allocation')
            max_need = banker_data.get('max_need')
            names = banker_data.get('names')
            if not total or not allocation or not max_need:
                errors.append('Dữ liệu Banker thiếu total, allocation hoặc max_need.')
            else:
                total = [int(x) for x in total]
                allocation = [[int(x) for x in row] for row in allocation]
                max_need = [[int(x) for x in row] for row in max_need]
                if names:
                    names = [str(x) for x in names]
                else:
                    names = [f"P{i}" for i in range(len(allocation))]
                request.session['custom_banker'] = {
                    'total': total,
                    'allocation': allocation,
                    'max_need': max_need,
                    'names': names
                }
                successes.append(f'Đã nhập dữ liệu Banker với {len(allocation)} tiến trình.')
                
        elif sim_type == 'memory':
            mem_data = data.get('memory', {})
            references = mem_data.get('references')
            frame_count = mem_data.get('frame_count', 3)
            if not references:
                errors.append('Dữ liệu bộ nhớ thiếu references.')
            else:
                references = [int(x) for x in references]
                frame_count = max(2, int(frame_count))
                request.session['custom_memory'] = {
                    'references': references,
                    'frame_count': frame_count
                }
                successes.append(f'Đã nhập {len(references)} tham chiếu trang bộ nhớ.')
                
        elif sim_type == 'synchronization':
            sync_data = data.get('synchronization', {})
            sequence = sync_data.get('sequence')
            buffer_size = sync_data.get('buffer_size', 3)
            producer_count = sync_data.get('producer_count', 2)
            consumer_count = sync_data.get('consumer_count', 2)
            if not sequence:
                errors.append('Dữ liệu đồng bộ hóa thiếu sequence.')
            else:
                sequence = [str(x).strip().upper()[0] for x in sequence if str(x).strip()]
                sequence = [x for x in sequence if x in ('P', 'C')]
                buffer_size = max(1, int(buffer_size))
                producer_count = max(1, int(producer_count))
                consumer_count = max(1, int(consumer_count))
                request.session['custom_sync'] = {
                    'sequence': sequence,
                    'buffer_size': buffer_size,
                    'producer_count': producer_count,
                    'consumer_count': consumer_count
                }
                successes.append(f'Đã nhập dữ liệu đồng bộ hóa với {len(sequence)} thao tác.')
                
        elif sim_type == 'machines':
            machines = data.get('machines', [])
            if not machines:
                errors.append('Danh sách máy trống.')
            else:
                valid_machines = []
                for idx, m in enumerate(machines):
                    name = str(m.get('name', '')).strip()
                    if not name:
                        continue
                    machine_type = str(m.get('machine_type', 'thuong')).strip().lower()
                    type_map = {
                        'thuong': 'thuong', 'máy thường': 'thuong', 'may thuong': 'thuong', 'regular': 'thuong',
                        'gaming': 'gaming', 'máy gaming': 'gaming', 'may gaming': 'gaming', 'game': 'gaming',
                        'vip': 'vip', 'phòng vip': 'vip', 'phong vip': 'vip',
                    }
                    machine_type = type_map.get(machine_type, 'thuong')
                    
                    status = str(m.get('status', 'trong')).strip().lower()
                    status_map = {
                        'trong': 'trong', 'trống': 'trong', 'available': 'trong', 'free': 'trong',
                        'dang_dung': 'dang_dung', 'đang dùng': 'dang_dung', 'dang dung': 'dang_dung', 'in use': 'dang_dung', 'busy': 'dang_dung',
                        'bao_tri': 'bao_tri', 'bảo trì': 'bao_tri', 'bao tri': 'bao_tri', 'maintenance': 'bao_tri',
                        'loi': 'loi', 'lỗi': 'loi', 'error': 'loi', 'fault': 'loi',
                        'khoa': 'khoa', 'khóa': 'khoa', 'khoá': 'khoa', 'locked': 'khoa',
                    }
                    status = status_map.get(status, 'trong')
                    
                    hourly_rate = m.get('hourly_rate', '10000')
                    try:
                        hourly_rate = int(float(str(hourly_rate).replace(',', '').replace('.', '')))
                        if hourly_rate < 0:
                            hourly_rate = 10000
                    except (ValueError, TypeError):
                        hourly_rate = 10000
                    
                    has_headset = str(m.get('has_headset', 'true')).strip().lower() in ('true', 'yes', '1', 'có', 'co')
                    has_account = str(m.get('has_account', 'false')).strip().lower() in ('true', 'yes', '1', 'có', 'co')
                    
                    ram_gb = m.get('ram_gb', '8')
                    try:
                        ram_gb = max(1, int(float(ram_gb)))
                    except (ValueError, TypeError):
                        ram_gb = 8
                    
                    note = str(m.get('note', '')).strip()
                    
                    valid_machines.append({
                        'name': name,
                        'machine_type': machine_type,
                        'status': status,
                        'hourly_rate': hourly_rate,
                        'has_headset': has_headset,
                        'has_account': has_account,
                        'ram_gb': ram_gb,
                        'note': note,
                    })
                
                if valid_machines:
                    request.session['custom_machines'] = valid_machines
                    successes.append(f'Đã nhập {len(valid_machines)} máy.')
                else:
                    errors.append('Không có máy hợp lệ để nhập.')
        else:
            errors.append(f'Loại mô phỏng "{sim_type}" không được hỗ trợ.')
    except (ValueError, TypeError) as e:
        errors.append(f'Lỗi kiểu dữ liệu: {str(e)}')
    except Exception as e:
        errors.append(f'Lỗi không xác định: {str(e)}')
    
    return (len(errors) == 0, successes + errors)


# ==============================================================
# API IMPORT — AJAX endpoint for real-time import with progress
# ==============================================================
@csrf_exempt
@login_required(login_url='app:login')
def api_import_simulation_data(request):

    """
    AJAX API endpoint for importing simulation data.
    Supports JSON, CSV, and Excel (.xlsx) files.
    Returns JSON response with validation results.
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)
    
    if 'file' not in request.FILES:
        return JsonResponse({'ok': False, 'error': 'Vui lòng chọn một file dữ liệu.'}, status=400)
    
    file = request.FILES['file']
    
    # --- File size validation (max 10MB) ---
    MAX_FILE_SIZE = 10 * 1024 * 1024
    if file.size > MAX_FILE_SIZE:
        return JsonResponse({
            'ok': False,
            'error': f'Dung lượng file quá lớn ({file.size / 1024 / 1024:.1f}MB). Tối đa 10MB.'
        }, status=400)
    
    # --- Empty file check ---
    if file.size == 0:
        return JsonResponse({'ok': False, 'error': 'File rỗng. Vui lòng chọn file có dữ liệu.'}, status=400)
    
    # --- Detect file type ---
    filename = file.name.lower()
    is_csv = filename.endswith('.csv')
    is_json = filename.endswith('.json')
    is_excel = filename.endswith('.xlsx') or filename.endswith('.xls')
    
    if not is_csv and not is_json and not is_excel:
        return JsonResponse({
            'ok': False,
            'error': 'Định dạng file không hỗ trợ. Chỉ chấp nhận file .json, .csv hoặc .xlsx.'
        }, status=400)
    
    # --- Parse file ---
    data = None
    errors = []
    warnings = []
    successes = []
    record_count = 0
    
    try:
        if is_json:
            data = json.load(file)
            record_count = _count_records(data)
            successes.append(f'Đọc thành công file JSON ({record_count} bản ghi).')
            
        elif is_csv:
            content = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
            if not rows:
                return JsonResponse({'ok': False, 'error': 'File CSV rỗng hoặc không có dữ liệu.'}, status=400)
            
            headers = reader.fieldnames or []
            headers_lower = [h.lower().strip() for h in headers]
            
            # Validate for null/empty values
            null_count = sum(1 for row in rows if all(v == '' or v is None for v in row.values()))
            if null_count > 0:
                warnings.append(f'Có {null_count} dòng dữ liệu trống/thiếu.')
            
            # Check for special characters
            special_char_count = 0
            for row in rows:
                for k, v in row.items():
                    if v and any(ord(c) > 127 and not c.isalpha() for c in str(v)):
                        special_char_count += 1
                        break
            if special_char_count > 0:
                warnings.append(f'Có {special_char_count} dòng chứa ký tự đặc biệt.')
            
            data, csv_errors = _parse_csv_data(rows, headers_lower)
            if csv_errors:
                errors.extend(csv_errors)
            else:
                record_count = _count_records(data)
                successes.append(f'Đọc thành công file CSV ({record_count} bản ghi).')
                
        elif is_excel:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                
                # Read headers from first row
                excel_headers = []
                for cell in ws[1]:
                    excel_headers.append(str(cell.value).strip() if cell.value else '')
                
                excel_headers_lower = [h.lower().strip() for h in excel_headers if h]
                
                # Read data rows
                excel_rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_data = {}
                    for i, val in enumerate(row):
                        if i < len(excel_headers) and excel_headers[i]:
                            row_data[excel_headers[i]] = str(val) if val is not None else ''
                    if any(v.strip() for v in row_data.values()):
                        excel_rows.append(row_data)
                
                wb.close()
                
                if not excel_rows:
                    return JsonResponse({'ok': False, 'error': 'File Excel rỗng hoặc không có dữ liệu.'}, status=400)
                
                data, excel_errors = _parse_csv_data(excel_rows, excel_headers_lower)
                if excel_errors:
                    errors.extend(excel_errors)
                else:
                    record_count = _count_records(data)
                    successes.append(f'Đọc thành công file Excel ({record_count} bản ghi).')
                    
            except ImportError:
                return JsonResponse({'ok': False, 'error': 'Thiếu thư viện openpyxl để đọc file Excel.'}, status=400)
            except Exception as e:
                return JsonResponse({'ok': False, 'error': f'Lỗi đọc file Excel: {str(e)}'}, status=400)
    
    except json.JSONDecodeError as e:
        return JsonResponse({'ok': False, 'error': f'Lỗi đọc file JSON: {str(e)}. Vui lòng kiểm tra định dạng JSON.'}, status=400)
    except csv.Error as e:
        return JsonResponse({'ok': False, 'error': f'Lỗi đọc file CSV: {str(e)}.'}, status=400)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'Lỗi xử lý file: {str(e)}'}, status=400)
    
    if data is None:
        return JsonResponse({'ok': False, 'error': 'Không thể đọc dữ liệu từ file.'}, status=400)
    
    # --- Auto-detect type ---
    sim_type = data.get('type')
    if not sim_type:
        if 'processes' in data:
            sim_type = 'scheduling'
        elif 'banker' in data:
            sim_type = 'banker'
        elif 'memory' in data:
            sim_type = 'memory'
        elif 'synchronization' in data:
            sim_type = 'synchronization'
        elif 'machines' in data:
            sim_type = 'machines'
        else:
            errors.append('Không thể xác định loại dữ liệu mô phỏng.')
    
    # --- Validate data based on type ---
    preview_data = None
    if sim_type == 'scheduling':
        processes = data.get('processes', [])
        if not processes:
            errors.append('Dữ liệu scheduling thiếu danh sách processes.')
        else:
            preview_data = _validate_scheduling_data(processes, errors, warnings, successes)
    elif sim_type == 'banker':
        preview_data = _validate_banker_data(data.get('banker', {}), errors, warnings, successes)
    elif sim_type == 'memory':
        preview_data = _validate_memory_data(data.get('memory', {}), errors, warnings, successes)
    elif sim_type == 'synchronization':
        preview_data = _validate_sync_data(data.get('synchronization', {}), errors, warnings, successes)
    elif sim_type == 'machines':
        preview_data = _validate_machines_data(data.get('machines', []), errors, warnings, successes)
    
    has_errors = len(errors) > 0
    
    # --- Save data to session if valid ---
    if not has_errors and data is not None:
        save_ok, save_messages = _save_imported_data(request, data, sim_type)
        if save_ok:
            successes.append('Dữ liệu đã được nhập vào hệ thống thành công.')
        else:
            errors.extend(save_messages)
            has_errors = True
    
    return JsonResponse({
        'ok': not has_errors,
        'type': sim_type,
        'record_count': record_count,
        'errors': errors,
        'warnings': warnings,
        'successes': successes,
        'preview': preview_data,
        'has_errors': has_errors,
    })


def _count_records(data):
    """Count records in parsed data."""
    if 'processes' in data:
        return len(data['processes'])
    if 'banker' in data:
        return len(data['banker'].get('allocation', []))
    if 'memory' in data:
        return len(data['memory'].get('references', []))
    if 'synchronization' in data:
        return len(data['synchronization'].get('sequence', []))
    if 'machines' in data:
        return len(data['machines'])
    return 0


def _parse_csv_data(rows, headers_lower):
    """Parse CSV/Excel rows into structured data."""
    errors = []
    
    if all(k in headers_lower for k in ['arrival', 'burst']):
        processes = []
        for i, row in enumerate(rows):
            try:
                arrival_val = row.get('arrival', '0')
                burst_val = row.get('burst', '10')
                if arrival_val is None or str(arrival_val).strip() == '':
                    errors.append(f'Dòng {i+2}: arrival time bị thiếu.')
                    continue
                if burst_val is None or str(burst_val).strip() == '':
                    errors.append(f'Dòng {i+2}: burst time bị thiếu.')
                    continue
                processes.append({
                    'pid': int(row.get('pid', i + 1)),
                    'name': str(row.get('name', f"P{row.get('pid', i + 1)}")),
                    'arrival': int(arrival_val),
                    'burst': int(burst_val),
                    'priority': int(row.get('priority', 0)),
                })
            except (ValueError, TypeError) as e:
                errors.append(f'Dòng {i+2}: Lỗi dữ liệu - {str(e)}')
        return {'type': 'scheduling', 'processes': processes, 'quantum': 20}, errors
        
    elif all(k in headers_lower for k in ['allocation', 'max_need']):
        total = []
        allocation = []
        max_need = []
        names = []
        for row in rows:
            try:
                alloc_str = str(row.get('allocation', '')).strip()
                max_str = str(row.get('max_need', '')).strip()
                if not alloc_str or not max_str:
                    errors.append('Dòng dữ liệu Banker bị thiếu allocation hoặc max_need.')
                    continue
                alloc_vals = [int(x.strip()) for x in alloc_str.replace('[','').replace(']','').split(',') if x.strip()]
                max_vals = [int(x.strip()) for x in max_str.replace('[','').replace(']','').split(',') if x.strip()]
                if not total and 'total' in headers_lower:
                    total_str = str(row.get('total', '')).strip()
                    total = [int(x.strip()) for x in total_str.replace('[','').replace(']','').split(',') if x.strip()]
                allocation.append(alloc_vals)
                max_need.append(max_vals)
                names.append(str(row.get('name', f"P{len(names)+1}")))
            except (ValueError, TypeError) as e:
                errors.append(f'Lỗi dữ liệu Banker: {str(e)}')
        return {
            'type': 'banker',
            'banker': {
                'total': total or [10, 5, 12],
                'allocation': allocation,
                'max_need': max_need,
                'names': names
            }
        }, errors
        
    elif 'references' in headers_lower:
        refs = []
        for row in rows:
            ref_str = str(row.get('references', '')).strip()
            if ref_str:
                refs.extend([int(x.strip()) for x in ref_str.split(',') if x.strip()])
        frame_count = int(rows[0].get('frame_count', 3)) if rows else 3
        return {
            'type': 'memory',
            'memory': {
                'references': refs,
                'frame_count': max(2, frame_count)
            }
        }, errors
        
    elif 'sequence' in headers_lower:
        seq = []
        for row in rows:
            s = str(row.get('sequence', '')).strip().upper()
            if s in ('P', 'C', 'PRODUCER', 'CONSUMER'):
                seq.append(s[0])
        buffer_size = int(rows[0].get('buffer_size', 3)) if rows else 3
        return {
            'type': 'synchronization',
            'synchronization': {
                'sequence': seq,
                'buffer_size': max(1, buffer_size),
                'producer_count': 3,
                'consumer_count': 2
            }
        }, errors
    
    elif all(k in headers_lower for k in ['name', 'machine_type']):
        """Parse machine data from CSV/Excel."""
        machines = []
        for i, row in enumerate(rows):
            try:
                name = str(row.get('name', '')).strip()
                if not name:
                    errors.append(f'Dòng {i+2}: Tên máy bị thiếu.')
                    continue
                machine_type = str(row.get('machine_type', 'thuong')).strip().lower()
                # Map type values
                type_map = {
                    'thuong': 'thuong', 'máy thường': 'thuong', 'may thuong': 'thuong', 'regular': 'thuong',
                    'gaming': 'gaming', 'máy gaming': 'gaming', 'may gaming': 'gaming', 'game': 'gaming',
                    'vip': 'vip', 'phòng vip': 'vip', 'phong vip': 'vip',
                }
                machine_type = type_map.get(machine_type, 'thuong')
                
                status = str(row.get('status', 'trong')).strip().lower()
                status_map = {
                    'trong': 'trong', 'trống': 'trong', 'available': 'trong', 'free': 'trong',
                    'dang_dung': 'dang_dung', 'đang dùng': 'dang_dung', 'dang dung': 'dang_dung', 'in use': 'dang_dung', 'busy': 'dang_dung',
                    'bao_tri': 'bao_tri', 'bảo trì': 'bao_tri', 'bao tri': 'bao_tri', 'maintenance': 'bao_tri',
                    'loi': 'loi', 'lỗi': 'loi', 'error': 'loi', 'fault': 'loi',
                    'khoa': 'khoa', 'khóa': 'khoa', 'khoá': 'khoa', 'locked': 'khoa',
                }
                status = status_map.get(status, 'trong')
                
                hourly_rate = row.get('hourly_rate', '10000')
                try:
                    hourly_rate = int(float(str(hourly_rate).replace(',', '').replace('.', '')))
                    if hourly_rate < 0:
                        hourly_rate = 10000
                except (ValueError, TypeError):
                    hourly_rate = 10000
                
                has_headset = str(row.get('has_headset', 'true')).strip().lower() in ('true', 'yes', '1', 'có', 'co')
                has_account = str(row.get('has_account', 'false')).strip().lower() in ('true', 'yes', '1', 'có', 'co')
                
                ram_gb = row.get('ram_gb', '8')
                try:
                    ram_gb = max(1, int(float(ram_gb)))
                except (ValueError, TypeError):
                    ram_gb = 8
                
                note = str(row.get('note', '')).strip()
                
                machines.append({
                    'name': name,
                    'machine_type': machine_type,
                    'status': status,
                    'hourly_rate': hourly_rate,
                    'has_headset': has_headset,
                    'has_account': has_account,
                    'ram_gb': ram_gb,
                    'note': note,
                })
            except Exception as e:
                errors.append(f'Dòng {i+2}: Lỗi xử lý - {str(e)}')
        
        return {'type': 'machines', 'machines': machines}, errors
    
    else:
        errors.append('Không thể xác định loại dữ liệu. Các cột cần có: arrival, burst (scheduling) hoặc allocation, max_need (banker) hoặc references (memory) hoặc sequence (sync) hoặc name, machine_type (machines).')
        return None, errors


def _validate_scheduling_data(processes, errors, warnings, successes):
    """Validate scheduling processes data."""
    valid_count = 0
    preview_rows = []
    seen_pids = set()
    
    for idx, p in enumerate(processes):
        pid = int(p.get('pid', idx + 1))
        name = str(p.get('name', f"P{idx+1}"))
        arrival = p.get('arrival', 0)
        burst = p.get('burst', 5)
        
        # Check for null values
        if arrival is None or burst is None:
            errors.append(f'Tiến trình {name}: arrival hoặc burst bị null.')
            continue
        
        # Check for negative values
        try:
            arrival = max(0, int(arrival))
            burst = max(1, int(burst))
        except (ValueError, TypeError):
            errors.append(f'Tiến trình {name}: arrival hoặc burst không phải số hợp lệ.')
            continue
        
        # Check for duplicates
        if pid in seen_pids:
            warnings.append(f'PID {pid} bị trùng lặp (tiến trình {name}).')
        seen_pids.add(pid)
        
        valid_count += 1
        preview_rows.append({
            'pid': pid,
            'name': name,
            'arrival': arrival,
            'burst': burst,
            'priority': int(p.get('priority', 3))
        })
    
    if valid_count == 0:
        errors.append('Không có tiến trình hợp lệ.')
    else:
        successes.append(f'{valid_count} tiến trình hợp lệ.')
    
    return {
        'columns': ['PID', 'Tên', 'Arrival', 'Burst', 'Priority'],
        'rows': preview_rows[:20],  # Preview first 20
        'total': valid_count
    }


def _validate_banker_data(banker_data, errors, warnings, successes):
    """Validate banker data."""
    total = banker_data.get('total')
    allocation = banker_data.get('allocation')
    max_need = banker_data.get('max_need')
    names = banker_data.get('names')
    
    if not total:
        errors.append('Thiếu tổng tài nguyên (total).')
        return None
    if not allocation:
        errors.append('Thiếu dữ liệu allocation.')
        return None
    if not max_need:
        errors.append('Thiếu dữ liệu max_need.')
        return None
    
    if len(allocation) != len(max_need):
        errors.append(f'Số lượng allocation ({len(allocation)}) và max_need ({len(max_need)}) không khớp.')
        return None
    
    # Check for null/negative values
    for i, alloc in enumerate(allocation):
        for j, val in enumerate(alloc):
            if val is None or val < 0:
                errors.append(f'Tiến trình {i}: allocation[{j}] không hợp lệ ({val}).')
    
    for i, need in enumerate(max_need):
        for j, val in enumerate(need):
            if val is None or val < 0:
                errors.append(f'Tiến trình {i}: max_need[{j}] không hợp lệ ({val}).')
    
    if not errors:
        successes.append(f'{len(allocation)} tiến trình, {len(total)} loại tài nguyên.')
    
    preview_rows = []
    for i in range(len(allocation)):
        preview_rows.append({
            'name': names[i] if names and i < len(names) else f'P{i}',
            'allocation': allocation[i],
            'max_need': max_need[i],
        })
    
    return {
        'columns': ['Tiến trình', 'Allocation', 'Max Need'],
        'rows': preview_rows,
        'total': len(allocation)
    }


def _validate_memory_data(mem_data, errors, warnings, successes):
    """Validate memory data."""
    references = mem_data.get('references')
    frame_count = mem_data.get('frame_count', 3)
    
    if not references:
        errors.append('Thiếu chuỗi tham chiếu trang (references).')
        return None
    
    # Check for null/negative values
    valid_refs = []
    for i, ref in enumerate(references):
        if ref is None:
            warnings.append(f'Tham chiếu {i} bị null, bỏ qua.')
            continue
        try:
            r = int(ref)
            if r < 0:
                warnings.append(f'Tham chiếu {i} âm ({r}), bỏ qua.')
                continue
            valid_refs.append(r)
        except (ValueError, TypeError):
            warnings.append(f'Tham chiếu {i} không hợp lệ ({ref}), bỏ qua.')
    
    if not valid_refs:
        errors.append('Không có tham chiếu trang hợp lệ.')
        return None
    
    successes.append(f'{len(valid_refs)} tham chiếu trang hợp lệ.')
    
    return {
        'columns': ['STT', 'Trang'],
        'rows': [{'index': i+1, 'page': ref} for i, ref in enumerate(valid_refs[:30])],
        'total': len(valid_refs),
        'frame_count': max(2, int(frame_count))
    }


def _validate_sync_data(sync_data, errors, warnings, successes):
    """Validate synchronization data."""
    sequence = sync_data.get('sequence')
    buffer_size = sync_data.get('buffer_size', 3)
    
    if not sequence:
        errors.append('Thiếu chuỗi thao tác (sequence).')
        return None
    
    valid_seq = [str(s).strip().upper()[0] for s in sequence if str(s).strip()]
    valid_seq = [s for s in valid_seq if s in ('P', 'C')]
    
    if not valid_seq:
        errors.append('Không có thao tác hợp lệ (P=Producer, C=Consumer).')
        return None
    
    p_count = valid_seq.count('P')
    c_count = valid_seq.count('C')
    successes.append(f'{len(valid_seq)} thao tác ({p_count} Producer, {c_count} Consumer).')
    
    return {
        'columns': ['STT', 'Thao tác'],
        'rows': [{'index': i+1, 'action': 'Producer' if s == 'P' else 'Consumer'} for i, s in enumerate(valid_seq[:30])],
        'total': len(valid_seq),
        'buffer_size': max(1, int(buffer_size))
    }


def _validate_machines_data(machines, errors, warnings, successes):
    """Validate machines data for import."""
    if not machines:
        errors.append('Danh sách máy trống.')
        return None
    
    valid_count = 0
    preview_rows = []
    seen_names = set()
    
    for idx, m in enumerate(machines):
        name = str(m.get('name', '')).strip()
        if not name:
            errors.append(f'Máy {idx+1}: Tên máy bị thiếu.')
            continue
        
        # Check for duplicates
        if name.lower() in seen_names:
            warnings.append(f'Máy "{name}" bị trùng lặp.')
        seen_names.add(name.lower())
        
        machine_type = str(m.get('machine_type', 'thuong')).strip().lower()
        type_map = {
            'thuong': 'thuong', 'máy thường': 'thuong', 'may thuong': 'thuong', 'regular': 'thuong',
            'gaming': 'gaming', 'máy gaming': 'gaming', 'may gaming': 'gaming', 'game': 'gaming',
            'vip': 'vip', 'phòng vip': 'vip', 'phong vip': 'vip',
        }
        machine_type = type_map.get(machine_type, 'thuong')
        
        status = str(m.get('status', 'trong')).strip().lower()
        status_map = {
            'trong': 'trong', 'trống': 'trong', 'available': 'trong', 'free': 'trong',
            'dang_dung': 'dang_dung', 'đang dùng': 'dang_dung', 'dang dung': 'dang_dung', 'in use': 'dang_dung', 'busy': 'dang_dung',
            'bao_tri': 'bao_tri', 'bảo trì': 'bao_tri', 'bao tri': 'bao_tri', 'maintenance': 'bao_tri',
            'loi': 'loi', 'lỗi': 'loi', 'error': 'loi', 'fault': 'loi',
            'khoa': 'khoa', 'khóa': 'khoa', 'khoá': 'khoa', 'locked': 'khoa',
        }
        status = status_map.get(status, 'trong')
        
        hourly_rate = m.get('hourly_rate', '10000')
        try:
            hourly_rate = int(float(str(hourly_rate).replace(',', '').replace('.', '')))
            if hourly_rate < 0:
                hourly_rate = 10000
        except (ValueError, TypeError):
            hourly_rate = 10000
        
        has_headset = str(m.get('has_headset', 'true')).strip().lower() in ('true', 'yes', '1', 'có', 'co')
        has_account = str(m.get('has_account', 'false')).strip().lower() in ('true', 'yes', '1', 'có', 'co')
        
        ram_gb = m.get('ram_gb', '8')
        try:
            ram_gb = max(1, int(float(ram_gb)))
        except (ValueError, TypeError):
            ram_gb = 8
        
        note = str(m.get('note', '')).strip()
        
        valid_count += 1
        preview_rows.append({
            'name': name,
            'machine_type': machine_type,
            'status': status,
            'hourly_rate': hourly_rate,
            'has_headset': has_headset,
            'has_account': has_account,
            'ram_gb': ram_gb,
            'note': note,
        })
    
    if valid_count == 0:
        errors.append('Không có máy hợp lệ để nhập.')
    else:
        successes.append(f'{valid_count} máy hợp lệ.')
    
    return {
        'columns': ['Tên máy', 'Loại', 'Trạng thái', 'Giá/giờ', 'Tai nghe', 'Tài khoản', 'RAM (GB)', 'Ghi chú'],
        'rows': preview_rows[:20],
        'total': valid_count
    }


@login_required(login_url='app:login')
def os_topics(request):
    """Trang tổng hợp các giải thuật và bài toán hệ điều hành phù hợp với đồ án."""
    core_topics = [
        {
            'title': 'Tiến trình',
            'role': 'Nền tảng của toàn bộ mô phỏng',
            'desc': 'Mỗi phiên chơi, khách hàng và hàng chờ được ánh xạ thành process với arrival time, burst time và resource vector.',
        },
        {
            'title': 'Lập lịch CPU',
            'role': 'Phần so sánh thuật toán trung tâm',
            'desc': 'Mô phỏng FCFS, SJF và Round Robin để quan sát waiting time, turnaround time, response time và CPU utilization.',
        },
        {
            'title': 'Bộ nhớ',
            'role': 'Đánh giá lỗi trang và hiệu quả cấp phát',
            'desc': 'Mô phỏng FIFO, LRU và OPT để so sánh số lỗi trang, hit rate và fault rate.',
        },
        {
            'title': 'Đồng bộ hóa',
            'role': 'Minh họa tương tranh giữa các tiến trình',
            'desc': 'Dùng mô hình Producer-Consumer để trình bày semaphore, vùng đệm hữu hạn và trạng thái chờ/được phục vụ.',
        },
        {
            'title': 'Banker / Deadlock',
            'role': 'Cơ chế quản lý tài nguyên an toàn',
            'desc': 'Kiểm tra trạng thái an toàn, mô phỏng request resource và giải thích điều kiện dẫn đến deadlock.',
        },
    ]

    supporting_topics = [
        {
            'title': 'Bakery Algorithm',
            'role': 'Mở rộng lý thuyết',
            'desc': 'Phù hợp đưa vào phần tham khảo khi cần nói về bảo đảm công bằng cho nhiều tiến trình vào cùng lúc.',
        },
        {
            'title': 'Test-and-Set',
            'role': 'Khóa nhị phân',
            'desc': 'Dùng để minh họa spinlock và cơ chế loại trừ tương tranh ở mức phần cứng.',
        },
        {
            'title': 'Swap',
            'role': 'Cơ chế đồng bộ đơn giản',
            'desc': 'Có thể nhắc như một cách tạo khóa đơn giản, nhưng không nên là trọng tâm báo cáo.',
        },
        {
            'title': 'Bounded-Buffer',
            'role': 'Producer - Consumer',
            'desc': 'Gắn trực tiếp với phần đồng bộ hóa và animation buffer đầy/trống trong app.',
        },
        {
            'title': 'Readers-Writers',
            'role': 'Truy cập dữ liệu chung',
            'desc': 'Phù hợp để liên hệ tới thống kê, báo cáo hoặc truy cập dữ liệu của hệ thống quản lý quán net.',
        },
        {
            'title': 'Dining Philosophers',
            'role': 'Bài toán deadlock kinh điển',
            'desc': 'Rất hay để giải thích deadlock, nhưng nên đặt ở phần tham khảo để tránh loãng trọng tâm.',
        },
    ]

    return render(request, 'app/os_topics.html', {
        'core_topics': core_topics,
        'supporting_topics': supporting_topics,
    })


# ==============================================================
# MÔ PHỎNG — LẬP LỊCH CPU
# ==============================================================
def _priority_from_customer(customer):
    member_type = (getattr(customer, 'member_type', '') or '').lower()
    if member_type == 'vip':
        return 1
    if member_type in {'thuong', 'regular'}:
        return 3
    return 5


def _priority_from_queue_item(queue_item):
    preferred_type = (getattr(queue_item, 'preferred_type', '') or '').lower()
    if preferred_type == 'vip':
        return 2
    if preferred_type in {'thuong', 'regular'}:
        return 4
    return 5


def _build_omdrrs_seed_processes(limit=500):
    """Tạo bộ tiến trình mặc định cho dashboard OMDRRS."""
    active_sessions = list(
        Session.objects.filter(status='dang_chay')
        .select_related('machine', 'customer')
        .order_by('start_time')[:limit]
    )
    waiting_queue = list(
        Queue.objects.filter(is_served=False)
        .order_by('arrived_at')[:limit]
    )

    processes = []
    pid = 1

    for sess in active_sessions:
        customer = getattr(sess, 'customer', None)
        priority = _priority_from_customer(customer) if customer else 4
        if getattr(sess, 'machine', None) and getattr(sess.machine, 'machine_type', '') == 'vip':
            priority = min(priority, 2)
        processes.append({
            'pid': pid,
            'name': f"P{pid}",
            'label': f"P{pid} - {sess.get_customer_name()}",
            'burst': int(getattr(sess, 'planned_minutes', 60) or 60),
            'arrival': int(sess.start_time.timestamp() // 60),
            'arrival_time': timezone.localtime(sess.start_time).strftime('%H:%M'),
            'arrival_datetime': timezone.localtime(sess.start_time).isoformat(),
            'priority': priority,
            'origin': 'live_session',
        })
        pid += 1

    for q in waiting_queue:
        processes.append({
            'pid': pid,
            'name': f"P{pid}",
            'label': f"P{pid} - {q.customer_name}",
            'burst': int(getattr(q, 'planned_minutes', 60) or 60),
            'arrival': int(q.arrived_at.timestamp() // 60),
            'arrival_time': timezone.localtime(q.arrived_at).strftime('%H:%M'),
            'arrival_datetime': timezone.localtime(q.arrived_at).isoformat(),
            'priority': _priority_from_queue_item(q),
            'origin': 'live_queue',
        })
        pid += 1

    for idx, proc in enumerate(processes, start=1):
        proc['pid'] = idx
        proc['name'] = f"P{idx}"
        proc['label'] = proc.get('label') or f"P{idx}"

    normalize_arrival_times(processes)
    return processes, active_sessions, waiting_queue


def _omdrrs_seed_for_frontend(seed_processes):
    """Chuẩn hóa seed JSON cho dashboard OMDRRS (pidLabel, burst, arrival, priority)."""
    normalized = []
    for proc in seed_processes:
        normalized.append({
            'uid': str(proc.get('pid', proc.get('name', '1'))),
            'pidLabel': proc.get('label') or proc.get('name') or f"P{proc.get('pid', 1)}",
            'burst': int(proc.get('burst', 10)),
            'arrival': int(proc.get('arrival', 0)),
            'priority': int(proc.get('priority', 3)),
        })
    return normalized


def _scheduling_bool_param(request, key, default=True):
    value = request.GET.get(key)
    if value is None:
        return default
    return value.lower() in ('on', 'true', '1', 'yes')


def _collect_manual_scheduling_processes(request, process_count):
    processes = []
    for index in range(1, process_count + 1):
        name = (request.GET.get(f'name_{index}') or f'P{index}').strip()
        try:
            arrival = int(request.GET.get(f'arrival_{index}', index - 1))
            burst = int(request.GET.get(f'burst_{index}', 5))
        except (TypeError, ValueError):
            continue
        processes.append({
            'pid': index,
            'name': name,
            'arrival': max(0, arrival),
            'burst': max(1, burst),
        })
    return processes


def _demo_scheduling_processes():
    return [
        {'pid': 1, 'name': 'P1', 'arrival': 0, 'burst': 8, 'priority': 2},
        {'pid': 2, 'name': 'P2', 'arrival': 0, 'burst': 3, 'priority': 1},
        {'pid': 3, 'name': 'P3', 'arrival': 0, 'burst': 9, 'priority': 4},
        {'pid': 4, 'name': 'P4', 'arrival': 0, 'burst': 6, 'priority': 1},
    ]


@login_required(login_url='app:login')
def scheduling_simulation(request):
    """Mô phỏng FCFS, SJF và Round Robin (giữ nguyên module cổ điển)."""
    expire_overdue_sessions()

    algorithm = (request.GET.get('algorithm') or 'fcfs').lower()
    if algorithm not in {'fcfs', 'sjf', 'rr', 'omdrrs'}:
        algorithm = 'fcfs'

    try:
        quantum = max(1, int(request.GET.get('quantum') or request.session.get('custom_quantum') or 30))
    except (TypeError, ValueError):
        quantum = 30

    data_mode = (request.GET.get('data_mode') or 'live').lower()
    include_active = _scheduling_bool_param(request, 'include_active', True)
    include_queue = _scheduling_bool_param(request, 'include_queue', True)

    try:
        process_count = max(1, min(500, int(request.GET.get('process_count', 5))))
    except (TypeError, ValueError):
        process_count = 5
    try:
        arrival_max = max(0, int(request.GET.get('arrival_max', 10)))
    except (TypeError, ValueError):
        arrival_max = 10
    try:
        burst_min = max(1, int(request.GET.get('burst_min', 5)))
    except (TypeError, ValueError):
        burst_min = 5
    try:
        burst_max = max(burst_min, int(request.GET.get('burst_max', 30)))
    except (TypeError, ValueError):
        burst_max = 30

    seed = (request.GET.get('seed') or '').strip()

    active_sessions = list(
        Session.objects.filter(status='dang_chay')
        .select_related('machine', 'customer')
        .order_by('start_time')[:500]
    )
    waiting_queue = list(
        Queue.objects.filter(is_served=False)
        .order_by('arrived_at')[:500]
    )

    source_label = 'Dữ liệu thật'
    source_description = 'Lấy từ phiên đang chạy và hàng chờ hiện có trong hệ thống quán net.'

    if data_mode == 'manual':
        processes = _collect_manual_scheduling_processes(request, process_count)
        source_label = 'Nhập tay'
        source_description = 'Bạn tự nhập arrival time và burst time cho từng tiến trình.'
    elif data_mode == 'random':
        processes = generate_random_processes(
            process_count,
            arrival_max=arrival_max,
            burst_min=burst_min,
            burst_max=burst_max,
            seed=seed or None,
        )
        source_label = 'Sinh ngẫu nhiên'
        source_description = 'Hệ thống tự sinh bộ tiến trình để demo các thuật toán lập lịch.'
    elif data_mode == 'demo':
        processes = _demo_scheduling_processes()
        source_label = 'Bo de mau'
        source_description = 'Bo 4 tien trinh P1-P4 dung de doi chieu ly thuyet FCFS, SJF va Round Robin.'
    elif data_mode == 'file':
        processes = request.session.get('custom_processes', [])
        source_label = 'Nhập từ file'
        source_description = 'Hệ thống sử dụng bộ tiến trình được tải lên từ file JSON.'
        if not processes:
            source_label = 'Du lieu file rong'
            source_description = 'Chua co du lieu file hop le, nen thuat toan khong co tien trinh de chay.'
    else:
        processes = []
        if include_active:
            processes.extend(sessions_to_processes(active_sessions))
        if include_queue:
            offset = len(processes)
            for idx, proc in enumerate(queue_to_processes(waiting_queue)):
                proc['pid'] = offset + idx + 1
                processes.append(proc)
        renumber_processes(processes)
        if not processes:
            source_label = 'Du lieu thuc te rong'
            source_description = 'Hien khong co phien dang chay hoac khach trong hang cho phu hop, nen thuat toan khong co tien trinh de chay.'

    normalize_arrival_times(processes)

    algorithm_names = {
        'fcfs': 'FCFS (First Come First Serve)',
        'sjf': 'SJF (Shortest Job First)',
        'rr': 'Round Robin',
        'omdrrs': 'OMDRRS',
    }
    algorithm_name = algorithm_names.get(algorithm, 'FCFS (First Come First Serve)')

    if algorithm == 'sjf':
        timeline = sjf(processes)
    elif algorithm == 'rr':
        timeline = round_robin(processes, quantum)
    elif algorithm == 'omdrrs':
        timeline = schedule_omdrrs(processes)
    else:
        timeline = fcfs(processes)

    stats = calc_stats(timeline, processes)
    events = build_scheduling_events(timeline, algorithm, quantum)

    total_ram = max(1, int(request.GET.get('total_ram', 1024)))
    mem_base = max(1, int(request.GET.get('mem_base', 64)))
    mem_per_minute = max(0, int(request.GET.get('mem_per_minute', 4)))
    cpu_memory = simulate_cpu_memory(
        timeline,
        total_ram=total_ram,
        mem_base=mem_base,
        mem_per_minute=mem_per_minute,
    )

    comparison = []
    best_wait = ''
    if processes:
        comparison = [
            {'key': 'fcfs', 'name': 'FCFS', 'stats': calc_stats(fcfs(processes), processes)},
            {'key': 'sjf', 'name': 'SJF', 'stats': calc_stats(sjf(processes), processes)},
            {'key': 'rr', 'name': 'Round Robin', 'stats': calc_stats(round_robin(processes, quantum), processes)},
            {'key': 'omdrrs', 'name': 'OMDRRS', 'stats': calc_stats(schedule_omdrrs(processes), processes)},
        ]
        best_wait = min(comparison, key=lambda row: row['stats']['avg_wait'])['key']

    timeline_payload = [
        {
            'pid': seg['pid'],
            'name': seg.get('name', f"P{seg['pid']}"),
            'customer_name': seg.get('customer_name', seg.get('name', '')),
            'arrival': seg.get('arrival', 0),
            'arrival_label': seg.get('arrival_label', f"t+{seg.get('arrival', 0)} phút"),
            'arrival_time': seg.get('arrival_time', ''),
            'arrival_display': seg.get('arrival_display') or (
                f"{seg.get('arrival_time')} (t+{seg.get('arrival', 0)} phút)"
                if seg.get('arrival_time')
                else f"t+{seg.get('arrival', 0)} phút"
            ),
            'burst': seg.get('burst', 0),
            'start': seg['start'],
            'end': seg['end'],
            'wait': seg.get('wait', 0),
            'tat': seg.get('tat', 0),
            'response': seg.get('response', seg.get('wait', 0)),
        }
        for seg in timeline
    ]

    # Làm giàu process data với task list để hiển thị chi tiết
    processes_enriched = []
    for proc in processes:
        enriched = dict(proc)
        enriched.setdefault('tasks', proc.get('tasks', []))
        enriched.setdefault('customer_name', proc.get('customer_name', proc.get('name', '')))
        enriched.setdefault('arrival_label', f"t+{proc.get('arrival', 0)} phút")
        enriched.setdefault(
            'arrival_display',
            f"{enriched.get('arrival_time')} ({enriched['arrival_label']})"
            if enriched.get('arrival_time')
            else enriched['arrival_label'],
        )
        processes_enriched.append(enriched)

    # Chuẩn hóa events JSON cho giao diện mới
    events_enriched = []
    for event in events:
        enriched = dict(event)
        enriched.setdefault('duration', enriched.get('end', 0) - enriched.get('start', 0))
        events_enriched.append(enriched)

    context = {
        'algorithm': algorithm,
        'algorithm_name': algorithm_name,
        'algorithm_description': ALGORITHM_DESCRIPTIONS.get(algorithm, ''),
        'quantum': quantum,
        'data_mode': data_mode,
        'include_active': include_active,
        'include_queue': include_queue,
        'process_count': process_count,
        'arrival_max': arrival_max,
        'burst_min': burst_min,
        'burst_max': burst_max,
        'seed': seed,
        'processes': processes_enriched,
        'timeline': timeline,
        'stats': stats,
        'comparison': comparison,
        'best_wait': best_wait,
        'processes_json': json.dumps(processes_enriched),
        'timeline_json': json.dumps(timeline_payload),
        'events_json': json.dumps(events_enriched),
        'source_label': source_label,
        'source_description': source_description,
        'cpu_memory': cpu_memory,
        'total_ram': total_ram,
        'mem_base': mem_base,
        'mem_per_minute': mem_per_minute,
    }
    return render(request, 'app/scheduling_sim.html', context)


@login_required(login_url='app:login')
def omdrrs_simulation(request):
    """
    Dashboard OMDRRS: mô phỏng trực quan theo thời gian thực và so sánh với Round Robin truyền thống.
    """
    expire_overdue_sessions()
    
    data_mode = request.GET.get('data_mode', 'live').lower()
    
    if data_mode == 'file':
        custom_procs = request.session.get('custom_processes', [])
        custom_machines = request.session.get('custom_machines', [])
        seed_processes = []
        
        # Ưu tiên 1: Dùng custom_processes (scheduling data)
        if custom_procs:
            for idx, p in enumerate(custom_procs):
                seed_processes.append({
                    'pid': p.get('pid', idx + 1),
                    'name': p.get('name', f"P{idx+1}"),
                    'label': f"P{p.get('pid', idx + 1)} - {p.get('name', f'P{idx+1}')}",
                    'burst': p.get('burst', 10),
                    'arrival': p.get('arrival', 0),
                    'priority': p.get('priority', 3),
                    'origin': 'file_import',
                })
            source_label = 'Nhập từ file (scheduling)'
            source_description = 'Bộ tiến trình được nhập từ file JSON.'
        
        # Ưu tiên 2: Dùng custom_machines (machines data) — chuyển máy thành tiến trình
        elif custom_machines:
            for idx, m in enumerate(custom_machines):
                # Map machine type to priority
                type_priority = {'thuong': 3, 'gaming': 2, 'vip': 1}
                priority = type_priority.get(m.get('machine_type', 'thuong'), 3)
                # Map status to burst time
                status_burst = {'trong': 15, 'dang_dung': 25, 'bao_tri': 5, 'loi': 3, 'khoa': 8}
                burst = status_burst.get(m.get('status', 'trong'), 15)
                # RAM affects burst
                ram_gb = int(m.get('ram_gb', 8))
                burst = max(3, burst + (ram_gb // 8) * 3)
                
                seed_processes.append({
                    'pid': idx + 1,
                    'name': m.get('name', f"PC{idx+1}"),
                    'label': f"PC{idx+1} - {m.get('name', f'PC{idx+1}')} ({m.get('machine_type', 'thuong')})",
                    'burst': burst,
                    'arrival': idx * 2,  # Staggered arrival
                    'priority': priority,
                    'origin': 'machine_import',
                })
            source_label = f'Nhập từ file (máy)'
            source_description = f'{len(custom_machines)} máy được chuyển đổi thành tiến trình để mô phỏng OMDRRS.'
        
        active_sessions = []
        waiting_queue = []
        
        if not seed_processes:
            active_sessions = []
            waiting_queue = []
            source_label = 'Du lieu file rong'
            source_description = 'Chua co du lieu file hop le, nen OMDRRS khong co tien trinh de chay.'
    else:
        seed_processes, active_sessions, waiting_queue = _build_omdrrs_seed_processes(limit=20)
        source_label = 'Dữ liệu thật'
        if active_sessions or waiting_queue:
            source_description = (
                'Kết hợp phiên đang chạy và hàng chờ hiện có trong hệ thống để tạo bộ tiến trình '
                'mang tính thực nghiệm cho OMDRRS.'
            )
        else:
            source_label = 'Du lieu thuc te rong'
            source_description = (
                'Hien khong co phien dang chay hoac khach trong hang cho, nen OMDRRS khong co tien trinh thuc te de chay.'
            )
    seed_frontend = _omdrrs_seed_for_frontend(seed_processes)
    if seed_processes:
        total_burst = sum(p['burst'] for p in seed_processes) or 1
        baseline_quantum = max(4, round(total_burst / len(seed_processes) * 0.6))
    else:
        baseline_quantum = 4

    context = {
        'seed_processes_json': json.dumps(seed_frontend),
        'source_label': source_label,
        'source_description': source_description,
        'baseline_quantum': baseline_quantum,
        'active_count': len(active_sessions),
        'queue_count': len(waiting_queue),
        'live_process_count': len(seed_frontend),
        'data_mode': data_mode,
    }
    return render(request, 'app/omdrrs_sim.html', context)


def _get_scheduling_processes_from_request(request):
    """Thu thập tiến trình dùng chung cho scheduling / compare."""
    try:
        quantum = max(1, int(request.GET.get('quantum') or request.session.get('custom_quantum') or 20))
    except (TypeError, ValueError):
        quantum = 20
    try:
        process_count = max(1, min(500, int(request.GET.get('process_count', 5))))
    except (TypeError, ValueError):
        process_count = 5
    try:
        arrival_max = max(0, int(request.GET.get('arrival_max', 10)))
    except (TypeError, ValueError):
        arrival_max = 10
    try:
        burst_min = max(1, int(request.GET.get('burst_min', 5)))
    except (TypeError, ValueError):
        burst_min = 5
    try:
        burst_max = max(burst_min, int(request.GET.get('burst_max', 30)))
    except (TypeError, ValueError):
        burst_max = 30

    data_mode = (request.GET.get('data_mode') or 'live').lower()
    include_active = _scheduling_bool_param(request, 'include_active', True)
    include_queue = _scheduling_bool_param(request, 'include_queue', True)
    seed = (request.GET.get('seed') or '').strip()
    try:
        q_episodes = max(20, min(300, int(request.GET.get('q_episodes', 100))))
    except (TypeError, ValueError):
        q_episodes = 100

    active_sessions = list(
        Session.objects.filter(status='dang_chay')
        .select_related('machine', 'customer')
        .order_by('start_time')[:500]
    )
    waiting_queue = list(
        Queue.objects.filter(is_served=False)
        .order_by('arrived_at')[:500]
    )

    source_label = 'Dữ liệu thật'
    source_description = 'Lấy từ phiên đang chạy và hàng chờ hiện có trong hệ thống quán net.'

    if data_mode == 'manual':
        processes = _collect_manual_scheduling_processes(request, process_count)
        source_label = 'Nhập tay'
        source_description = 'Bạn tự nhập arrival time và burst time cho từng tiến trình.'
    elif data_mode == 'random':
        processes = generate_random_processes(
            process_count,
            arrival_max=arrival_max,
            burst_min=burst_min,
            burst_max=burst_max,
            seed=seed or None,
        )
        source_label = 'Sinh ngẫu nhiên'
        source_description = 'Hệ thống tự sinh bộ tiến trình để so sánh thuật toán.'
    elif data_mode == 'demo':
        processes = _demo_scheduling_processes()
        source_label = 'Bo de mau'
        source_description = 'Bo 4 tien trinh P1-P4 dung de doi chieu ly thuyet FCFS, SJF va Round Robin.'
    elif data_mode == 'file':
        processes = request.session.get('custom_processes', [])
        source_label = 'Nhập từ file'
        source_description = 'Hệ thống sử dụng bộ tiến trình được tải lên từ file JSON.'
        if not processes:
            source_label = 'Du lieu file rong'
            source_description = 'Chua co du lieu file hop le, nen thuat toan khong co tien trinh de chay.'
    else:
        processes = []
        if include_active:
            processes.extend(sessions_to_processes(active_sessions))
        if include_queue:
            offset = len(processes)
            for idx, proc in enumerate(queue_to_processes(waiting_queue)):
                proc['pid'] = offset + idx + 1
                processes.append(proc)
        renumber_processes(processes)
        if not processes:
            source_label = 'Du lieu thuc te rong'
            source_description = 'Hien khong co phien dang chay hoac khach trong hang cho phu hop, nen thuat toan khong co tien trinh de chay.'

    normalize_arrival_times(processes)

    for proc in processes:
        proc.setdefault('priority', 3)

    return {
        'processes': processes,
        'quantum': quantum,
        'q_episodes': q_episodes,
        'seed': seed,
        'data_mode': data_mode,
        'process_count': process_count,
        'arrival_max': arrival_max,
        'burst_min': burst_min,
        'burst_max': burst_max,
        'include_active': include_active,
        'include_queue': include_queue,
        'source_label': source_label,
        'source_description': source_description,
    }


@login_required(login_url='app:login')
def scheduling_compare(request):
    """
    So sánh song song FCFS, SJF, RR, OMDRRS và Q-Learning với animation đồng bộ.
    """
    expire_overdue_sessions()
    cfg = _get_scheduling_processes_from_request(request)
    processes = cfg['processes']
    run_compare = request.GET.get('run') == '1'

    comparison_results = []
    best_wait_key = ''
    if run_compare and processes:
        comparison_results = run_all_schedulers(
            processes,
            quantum=cfg['quantum'],
            q_episodes=cfg['q_episodes'],
            seed=cfg['seed'] or None,
        )
        best_wait_key = min(comparison_results, key=lambda r: r['stats']['avg_wait'])['key']

    context = {
        **cfg,
        'run_compare': run_compare,
        'comparison_results': comparison_results,
        'comparison_json': json.dumps(comparison_results),
        'processes_json': json.dumps(processes),
        'best_wait_key': best_wait_key,
    }
    return render(request, 'app/scheduling_compare.html', context)


# ==============================================================
# MÔ PHỎNG — BANKER ALGORITHM
# ==============================================================
@login_required(login_url='app:login')
def banker_simulation(request):
    """
    Mô phỏng Banker Algorithm - kiểm tra tính an toàn khi cấp phát tài nguyên
    """
    expire_overdue_sessions()
    scenario = request.GET.get('scenario') or request.POST.get('scenario') or 'live'
    policy = request.GET.get('policy') or request.POST.get('policy') or 'min_resource_loss'

    scenario_data = {
        'safe_case': {
            'label': 'Case Mau An Toan',
            'total': [10, 5, 12],
            'allocation': [[1, 1, 2], [2, 0, 1], [1, 1, 0]],
            'max_need': [[3, 2, 3], [3, 2, 2], [2, 2, 1]],
            'names': ['P0 - Khach A', 'P1 - Khach B', 'P2 - Khach C'],
        },
        'unsafe_case': {
            'label': 'Case Mau Khong An Toan',
            'total': [6, 3, 6],
            'allocation': [[2, 1, 2], [2, 1, 1], [1, 0, 2]],
            'max_need': [[3, 2, 3], [3, 2, 2], [2, 2, 3]],
            'names': ['P0 - Khach X', 'P1 - Khach Y', 'P2 - Khach Z'],
        },
        'deadlock_case': {
            'label': 'Case Mau Deadlock',
            'total': [4, 2, 4],
            'allocation': [[2, 1, 1], [1, 1, 2], [1, 0, 1]],
            'max_need': [[3, 2, 2], [2, 2, 3], [2, 1, 3]],
            'names': ['P0 - May 01', 'P1 - May 02', 'P2 - May 03'],
        },
        'ten_machines_case': {
            'label': 'Case Banker 10 Máy',
            'total': [10, 5, 12],
            'allocation': [
                [0, 0, 1],
                [1, 0, 0],
                [2, 1, 1],
                [1, 0, 1],
                [0, 0, 2],
                [1, 1, 0],
                [1, 0, 0],
                [0, 1, 1],
                [1, 0, 1],
                [0, 0, 1],
            ],
            'max_need': [
                [1, 1, 1],
                [1, 1, 0],
                [2, 1, 2],
                [1, 1, 2],
                [1, 1, 2],
                [1, 1, 1],
                [2, 0, 1],
                [1, 1, 1],
                [2, 1, 2],
                [1, 0, 1],
            ],
            'names': [
                'P0 - Máy 01', 'P1 - Máy 02', 'P2 - Máy 03', 'P3 - Máy 04', 'P4 - Máy 05',
                'P5 - Máy 06', 'P6 - Máy 07', 'P7 - Máy 08', 'P8 - Máy 09', 'P9 - Máy 10',
            ],
        },
    }

    active_sessions = list(Session.objects.filter(status='dang_chay').select_related('machine', 'customer'))
    using_live = scenario == 'live'

    if scenario == 'file':
        custom_banker = request.session.get('custom_banker', {})
        total = custom_banker.get('total', [10, 5, 12])
        allocation = custom_banker.get('allocation', [[1, 1, 2], [2, 0, 1], [1, 1, 0]])
        max_need = custom_banker.get('max_need', [[3, 2, 3], [3, 2, 2], [2, 2, 1]])
        names = custom_banker.get('names', [f"P{i}" for i in range(len(allocation))])
        need = [
            [max_need[i][j] - allocation[i][j] for j in range(len(total))]
            for i in range(len(allocation))
        ]
        used = [sum(allocation[i][j] for i in range(len(allocation))) for j in range(len(total))]
        available = [max(0, total[j] - used[j]) for j in range(len(total))]
        banker_data = {
            'n': len(allocation),
            'm': len(total),
            'resource_names': ['Tai Nghe', 'Tai Khoan Game', 'RAM (GB)'],
            'total': total,
            'available': available,
            'allocation': allocation,
            'max_need': max_need,
            'need': need,
        }
        session_names = names
        active_sessions = []
    elif using_live:
        banker_data = build_banker_data(active_sessions)
        session_names = [f"S{s.id} - {s.get_customer_name()} ({s.machine.name})" for s in active_sessions]
    else:
        case = scenario_data.get(scenario, scenario_data['safe_case'])
        total = case['total']
        allocation = case['allocation']
        max_need = case['max_need']
        need = [
            [max_need[i][j] - allocation[i][j] for j in range(len(total))]
            for i in range(len(allocation))
        ]
        used = [sum(allocation[i][j] for i in range(len(allocation))) for j in range(len(total))]
        available = [max(0, total[j] - used[j]) for j in range(len(total))]
        banker_data = {
            'n': len(allocation),
            'm': len(total),
            'resource_names': ['Tai Nghe', 'Tai Khoan Game', 'RAM (GB)'],
            'total': total,
            'available': available,
            'allocation': allocation,
            'max_need': max_need,
            'need': need,
        }
        session_names = case['names']
        active_sessions = []

    safety_result = banker_safety(
        banker_data['n'],
        banker_data['available'],
        banker_data['allocation'],
        banker_data['need'],
        session_names=session_names,
    ) if banker_data['n'] > 0 else {'safe': True, 'safe_sequence': [], 'trace': [], 'deadlock_info': {'is_deadlock': False, 'blocked_processes': [], 'reason': '', 'solution': '', 'deadlock_log': []}}

    request_result = None
    if request.method == 'POST' and banker_data['n'] > 0:
        try:
            session_index = int(request.POST.get('session_index', '0'))
            req_headset = int(request.POST.get('req_headset', '0'))
            req_account = int(request.POST.get('req_account', '0'))
            req_ram = int(request.POST.get('req_ram', '0'))
            req_vector = [max(0, req_headset), max(0, req_account), max(0, req_ram)]
            if 0 <= session_index < banker_data['n']:
                request_result = banker_request(req_vector, session_index, banker_data)
                request_result['session_index'] = session_index
                request_result['request_vector'] = req_vector
        except (TypeError, ValueError):
            request_result = {'can_grant': False, 'reason': 'Du lieu request khong hop le.', 'safe_after': False, 'sequence': []}

    deadlock_resolution = None
    if safety_result['deadlock_info']['is_deadlock']:
        blocked = safety_result['deadlock_info']['blocked_processes']
        if blocked:
            if policy == 'priority_based' and using_live:
                def priority_score(i):
                    if i >= len(active_sessions) or not active_sessions[i].customer:
                        return (0, sum(banker_data['allocation'][i]))
                    member = (active_sessions[i].customer.member_type or '').lower()
                    is_vip = member in ['vip', 'kim_cuong', 'platinum']
                    return (1 if is_vip else 0, sum(banker_data['allocation'][i]))
                victim = min(blocked, key=priority_score)
            elif policy == 'min_waiting_impact':
                victim = min(blocked, key=lambda i: sum(banker_data['need'][i]))
            else:
                victim = min(blocked, key=lambda i: sum(banker_data['allocation'][i]))

            freed = banker_data['allocation'][victim]
            avail_new = [banker_data['available'][j] + freed[j] for j in range(banker_data['m'])]
            alloc_new = [row for idx, row in enumerate(banker_data['allocation']) if idx != victim]
            need_new = [row for idx, row in enumerate(banker_data['need']) if idx != victim]
            names_new = [nm for idx, nm in enumerate(session_names) if idx != victim]
            after = banker_safety(len(alloc_new), avail_new, alloc_new, need_new, names_new) if alloc_new else {'safe': True, 'safe_sequence': [], 'trace': []}
            deadlock_resolution = {
                'policy': policy,
                'victim_index': victim,
                'victim_name': session_names[victim] if victim < len(session_names) else f'P{victim}',
                'freed': freed,
                'safe_after': after['safe'],
                'sequence_after': after['safe_sequence'],
            }

    is_safe = safety_result['safe']
    safe_sequence = safety_result['safe_sequence']
    trace = safety_result['trace']
    deadlock_info = safety_result['deadlock_info']

    total_resources = {'headset': banker_data['total'][0], 'account': banker_data['total'][1], 'ram_gb': banker_data['total'][2]}
    used_resources = {'headset': total_resources['headset'] - banker_data['available'][0], 'account': total_resources['account'] - banker_data['available'][1], 'ram_gb': total_resources['ram_gb'] - banker_data['available'][2]}
    available_resources = {'headset': banker_data['available'][0], 'account': banker_data['available'][1], 'ram_gb': banker_data['available'][2]}

    banker_available_rows = [
        {'name': 'Tai Nghe', 'value': banker_data['available'][0]},
        {'name': 'Tai Khoan Game', 'value': banker_data['available'][1]},
        {'name': 'RAM (GB)', 'value': banker_data['available'][2]},
    ]
    banker_session_rows = []
    for i in range(banker_data['n']):
        session_obj = active_sessions[i] if using_live and i < len(active_sessions) else None
        banker_session_rows.append({
            'session': session_obj if session_obj else type('Obj', (), {'id': i}),
            'display_name': session_names[i] if i < len(session_names) else f'P{i}',
            'allocation': banker_data['allocation'][i],
            'max_need': banker_data['max_need'][i],
            'need': banker_data['need'][i],
            'index': i,
        })

    state_level = 'safe' if is_safe else ('deadlock' if deadlock_info.get('is_deadlock') else 'unsafe')

    context = {
        'total_resources': total_resources,
        'used_resources': used_resources,
        'available_resources': available_resources,
        'active_sessions': active_sessions,
        'is_safe': is_safe,
        'safe_sequence': safe_sequence,
        'trace': trace,
        'deadlock_info': deadlock_info,
        'banker_available_rows': banker_available_rows,
        'banker_session_rows': banker_session_rows,
        'banker_data': banker_data,
        'scenario': scenario,
        'state_level': state_level,
        'request_result': request_result,
        'deadlock_resolution': deadlock_resolution,
        'policy': policy,
        'using_live': using_live,
    }
    return render(request, 'app/banker_sim.html', context)


# ==============================================================
# MÔ PHỎNG — DBDAA (Dynamic Banker's Deadlock Avoidance Algorithm)
# ==============================================================
@login_required(login_url='app:login')
def dbdaa_simulation(request):
    """
    Mô phỏng DBDAA - Dynamic Banker's Deadlock Avoidance Algorithm
    Cải tiến từ Banker truyền thống với SRQ, Fast-track Check và Primary Unsafe Sequence
    """
    expire_overdue_sessions()
    scenario = request.GET.get('scenario') or request.POST.get('scenario') or 'live'

    scenario_data = {
        'safe_case': {
            'label': 'Case Mau An Toan',
            'total': [10, 5, 12],
            'allocation': [[1, 1, 2], [2, 0, 1], [1, 1, 0]],
            'max_need': [[3, 2, 3], [3, 2, 2], [2, 2, 1]],
            'names': ['P0 - Khach A', 'P1 - Khach B', 'P2 - Khach C'],
        },
        'unsafe_case': {
            'label': 'Case Mau Khong An Toan',
            'total': [6, 3, 6],
            'allocation': [[2, 1, 2], [2, 1, 1], [1, 0, 2]],
            'max_need': [[3, 2, 3], [3, 2, 2], [2, 2, 3]],
            'names': ['P0 - Khach X', 'P1 - Khach Y', 'P2 - Khach Z'],
        },
        'deadlock_case': {
            'label': 'Case Mau Deadlock',
            'total': [4, 2, 4],
            'allocation': [[2, 1, 1], [1, 1, 2], [1, 0, 1]],
            'max_need': [[3, 2, 2], [2, 2, 3], [2, 1, 3]],
            'names': ['P0 - May 01', 'P1 - May 02', 'P2 - May 03'],
        },
    }

    active_sessions = list(Session.objects.filter(status='dang_chay').select_related('machine', 'customer'))
    using_live = scenario == 'live'

    if scenario == 'file':
        custom_banker = request.session.get('custom_banker', {})
        total = custom_banker.get('total', [10, 5, 12])
        allocation = custom_banker.get('allocation', [[1, 1, 2], [2, 0, 1], [1, 1, 0]])
        max_need = custom_banker.get('max_need', [[3, 2, 3], [3, 2, 2], [2, 2, 1]])
        names = custom_banker.get('names', [f"P{i}" for i in range(len(allocation))])
        need = [
            [max_need[i][j] - allocation[i][j] for j in range(len(total))]
            for i in range(len(allocation))
        ]
        used = [sum(allocation[i][j] for i in range(len(allocation))) for j in range(len(total))]
        available = [max(0, total[j] - used[j]) for j in range(len(total))]
        banker_data = {
            'n': len(allocation),
            'm': len(total),
            'resource_names': ['Tai Nghe', 'Tai Khoan Game', 'RAM (GB)'],
            'total': total,
            'available': available,
            'allocation': allocation,
            'max_need': max_need,
            'need': need,
        }
        session_names = names
        active_sessions = []
    elif using_live:
        banker_data = build_banker_data(active_sessions)
        session_names = [f"S{s.id} - {s.get_customer_name()} ({s.machine.name})" for s in active_sessions]
    else:
        case = scenario_data.get(scenario, scenario_data['safe_case'])
        total = case['total']
        allocation = case['allocation']
        max_need = case['max_need']
        need = [
            [max_need[i][j] - allocation[i][j] for j in range(len(total))]
            for i in range(len(allocation))
        ]
        used = [sum(allocation[i][j] for i in range(len(allocation))) for j in range(len(total))]
        available = [max(0, total[j] - used[j]) for j in range(len(total))]
        banker_data = {
            'n': len(allocation),
            'm': len(total),
            'resource_names': ['Tai Nghe', 'Tai Khoan Game', 'RAM (GB)'],
            'total': total,
            'available': available,
            'allocation': allocation,
            'max_need': max_need,
            'need': need,
        }
        session_names = case['names']
        active_sessions = []

    safety_result = dbdaa_safety(
        banker_data['n'],
        banker_data['available'],
        banker_data['allocation'],
        banker_data['need'],
        session_names=session_names,
    ) if banker_data['n'] > 0 else {'safe': True, 'safe_sequence': [], 'trace': [], 'primary_unsafe': [], 'deadlock_info': {'is_deadlock': False, 'blocked_processes': [], 'reason': '', 'solution': '', 'deadlock_log': []}, 'performance': {'total_comparisons': 0, 'fast_track_skips': 0, 'srq_reorderings': 0}}

    request_result = None
    if request.method == 'POST' and banker_data['n'] > 0:
        try:
            session_index = int(request.POST.get('session_index', '0'))
            req_headset = int(request.POST.get('req_headset', '0'))
            req_account = int(request.POST.get('req_account', '0'))
            req_ram = int(request.POST.get('req_ram', '0'))
            req_vector = [max(0, req_headset), max(0, req_account), max(0, req_ram)]
            if 0 <= session_index < banker_data['n']:
                request_result = dbdaa_request(req_vector, session_index, banker_data)
                request_result['session_index'] = session_index
                request_result['request_vector'] = req_vector
        except (TypeError, ValueError):
            request_result = {'can_grant': False, 'reason': 'Du lieu request khong hop le.', 'safe_after': False, 'sequence': [], 'performance': {}}

    is_safe = safety_result['safe']
    safe_sequence = safety_result['safe_sequence']
    trace = safety_result['trace']
    primary_unsafe = safety_result['primary_unsafe']
    deadlock_info = safety_result['deadlock_info']
    performance = safety_result['performance']

    total_resources = {'headset': banker_data['total'][0], 'account': banker_data['total'][1], 'ram_gb': banker_data['total'][2]}
    used_resources = {'headset': total_resources['headset'] - banker_data['available'][0], 'account': total_resources['account'] - banker_data['available'][1], 'ram_gb': total_resources['ram_gb'] - banker_data['available'][2]}
    available_resources = {'headset': banker_data['available'][0], 'account': banker_data['available'][1], 'ram_gb': banker_data['available'][2]}

    banker_session_rows = []
    for i in range(banker_data['n']):
        session_obj = active_sessions[i] if using_live and i < len(active_sessions) else None
        banker_session_rows.append({
            'session_id': session_obj.id if session_obj else i,
            'display_name': session_names[i] if i < len(session_names) else f'P{i}',
            'allocation': banker_data['allocation'][i],
            'max_need': banker_data['max_need'][i],
            'need': banker_data['need'][i],
            'index': i,
            'max_need_total': sum(banker_data['need'][i]),
        })

    state_level = 'safe' if is_safe else ('deadlock' if deadlock_info.get('is_deadlock') else 'unsafe')

    context = {
        'total_resources': total_resources,
        'used_resources': used_resources,
        'available_resources': available_resources,
        'active_sessions': active_sessions,
        'is_safe': is_safe,
        'safe_sequence': safe_sequence,
        'trace': trace,
        'primary_unsafe': primary_unsafe,
        'deadlock_info': deadlock_info,
        'performance': performance,
        'banker_session_rows': banker_session_rows,
        'scenario': scenario,
        'state_level': state_level,
        'request_result': request_result,
        'using_live': using_live,
    }
    return render(request, 'app/dbdaa_sim.html', context)


# ==============================================================
# SO SÁNH BANKER VÀ DBDAA
# ==============================================================
@login_required(login_url='app:login')
def banker_dbdaa_compare(request):
    """
    So sánh giữa Banker truyền thống và DBDAA
    """
    expire_overdue_sessions()
    scenario = request.GET.get('scenario') or 'live'

    scenario_data = {
        'safe_case': {
            'label': 'Case Mau An Toan',
            'total': [10, 5, 12],
            'allocation': [[1, 1, 2], [2, 0, 1], [1, 1, 0]],
            'max_need': [[3, 2, 3], [3, 2, 2], [2, 2, 1]],
            'names': ['P0 - Khach A', 'P1 - Khach B', 'P2 - Khach C'],
        },
        'unsafe_case': {
            'label': 'Case Mau Khong An Toan',
            'total': [6, 3, 6],
            'allocation': [[2, 1, 2], [2, 1, 1], [1, 0, 2]],
            'max_need': [[3, 2, 3], [3, 2, 2], [2, 2, 3]],
            'names': ['P0 - Khach X', 'P1 - Khach Y', 'P2 - Khach Z'],
        },
        'deadlock_case': {
            'label': 'Case Mau Deadlock',
            'total': [4, 2, 4],
            'allocation': [[2, 1, 1], [1, 1, 2], [1, 0, 1]],
            'max_need': [[3, 2, 2], [2, 2, 3], [2, 1, 3]],
            'names': ['P0 - May 01', 'P1 - May 02', 'P2 - May 03'],
        },
    }

    active_sessions = list(Session.objects.filter(status='dang_chay').select_related('machine', 'customer'))
    using_live = scenario == 'live'
    scenario_label = 'Du lieu thuc te'

    if scenario == 'file':
        custom_banker = request.session.get('custom_banker', {})
        total = custom_banker.get('total', [10, 5, 12])
        allocation = custom_banker.get('allocation', [[1, 1, 2], [2, 0, 1], [1, 1, 0]])
        max_need = custom_banker.get('max_need', [[3, 2, 3], [3, 2, 2], [2, 2, 1]])
        names = custom_banker.get('names', [f"P{i}" for i in range(len(allocation))])
        need = [
            [max_need[i][j] - allocation[i][j] for j in range(len(total))]
            for i in range(len(allocation))
        ]
        used = [sum(allocation[i][j] for i in range(len(allocation))) for j in range(len(total))]
        available = [max(0, total[j] - used[j]) for j in range(len(total))]
        banker_data = {
            'n': len(allocation),
            'm': len(total),
            'resource_names': ['Tai Nghe', 'Tai Khoan Game', 'RAM (GB)'],
            'total': total,
            'available': available,
            'allocation': allocation,
            'max_need': max_need,
            'need': need,
        }
        session_names = names
        scenario_label = 'Nhap tu file'
    elif using_live:
        banker_data = build_banker_data(active_sessions)
        session_names = [f"S{s.id} - {s.get_customer_name()} ({s.machine.name})" for s in active_sessions]
    else:
        case = scenario_data.get(scenario, scenario_data['safe_case'])
        total = case['total']
        allocation = case['allocation']
        max_need = case['max_need']
        need = [
            [max_need[i][j] - allocation[i][j] for j in range(len(total))]
            for i in range(len(allocation))
        ]
        used = [sum(allocation[i][j] for i in range(len(allocation))) for j in range(len(total))]
        available = [max(0, total[j] - used[j]) for j in range(len(total))]
        banker_data = {
            'n': len(allocation),
            'm': len(total),
            'resource_names': ['Tai Nghe', 'Tai Khoan Game', 'RAM (GB)'],
            'total': total,
            'available': available,
            'allocation': allocation,
            'max_need': max_need,
            'need': need,
        }
        session_names = case['names']
        scenario_label = case['label']

    if banker_data['n'] > 0:
        banker_result = banker_safety(
            banker_data['n'],
            banker_data['available'],
            banker_data['allocation'],
            banker_data['need'],
            session_names=session_names,
        )
        dbdaa_result = dbdaa_safety(
            banker_data['n'],
            banker_data['available'],
            banker_data['allocation'],
            banker_data['need'],
            session_names=session_names,
        )
    else:
        banker_result = {'safe': True, 'safe_sequence': [], 'trace': []}
        dbdaa_result = {
            'safe': True,
            'safe_sequence': [],
            'trace': [],
            'primary_unsafe': [],
            'performance': {'total_comparisons': 0, 'fast_track_skips': 0, 'srq_reorderings': 0},
        }

    # Tính toán số phép so sánh cho Banker (ước tính)
    banker_comparisons = banker_data['n'] * banker_data['n'] * banker_data['m']
    
    # So sánh hiệu suất
    comparison = {
        'banker_safe': banker_result['safe'],
        'dbdaa_safe': dbdaa_result['safe'],
        'banker_sequence': banker_result['safe_sequence'],
        'dbdaa_sequence': dbdaa_result['safe_sequence'],
        'banker_trace': banker_result['trace'],
        'dbdaa_trace': dbdaa_result['trace'],
        'banker_comparisons': banker_comparisons,
        'dbdaa_comparisons': dbdaa_result['performance']['total_comparisons'],
        'dbdaa_fast_track_skips': dbdaa_result['performance']['fast_track_skips'],
        'dbdaa_srq_reorderings': dbdaa_result['performance']['srq_reorderings'],
        'dbdaa_primary_unsafe': dbdaa_result['primary_unsafe'],
        'improvement_ratio': round(banker_comparisons / max(1, dbdaa_result['performance']['total_comparisons']), 2),
        'skip_ratio': round(dbdaa_result['performance']['fast_track_skips'] / max(1, dbdaa_result['performance']['total_comparisons']) * 100, 1),
    }

    context = {
        'scenario': scenario,
        'scenario_label': scenario_label,
        'banker_data': banker_data,
        'session_names': session_names,
        'comparison': comparison,
    }
    return render(request, 'app/banker_dbdaa_compare.html', context)


# ==============================================================
# MÔ PHỎNG — BỘ NHỚ
# ==============================================================
@login_required(login_url='app:login')
def memory_simulation(request):
    expire_overdue_sessions()

    data_mode = request.GET.get('data_mode', 'live').lower()
    algorithm = request.GET.get('algorithm', 'lru').lower()
    frame_count = max(2, int(request.GET.get('frame_count', 3)))
    ref_length = max(6, int(request.GET.get('ref_length', 18)))
    seed = request.GET.get('seed', '').strip()

    sessions = list(Session.objects.filter(status='dang_chay').select_related('machine', 'customer').order_by('start_time')[:8])
    queue_items = list(Queue.objects.filter(is_served=False).order_by('arrived_at')[:8])

    reference_source_label = 'Dữ liệu thật'
    reference_source_description = 'Lấy từ phiên đang chạy và hàng chờ hiện có trong hệ thống.'
    manual_reference_raw = request.GET.get('references', '')

    if data_mode == 'manual':
        refs = parse_reference_string(manual_reference_raw, limit=ref_length)
        reference_source_label = 'Nhập tay'
        reference_source_description = 'Bạn tự nhập chuỗi trang để so sánh FIFO, LRU và OPT.'
    elif data_mode == 'file':
        custom_memory = request.session.get('custom_memory', {})
        refs = custom_memory.get('references', [1, 2, 3, 4, 1, 2, 5, 1, 2, 3, 4, 5])
        frame_count = max(2, int(request.GET.get('frame_count') or custom_memory.get('frame_count', 3)))
        ref_length = len(refs)
        reference_source_label = 'Nhập từ file'
        reference_source_description = 'Chuỗi trang truy cập được nhập từ file tải lên.'
    else:
        refs = build_memory_reference_stream(
            data_mode,
            sessions=sessions,
            queue_items=queue_items,
            length=ref_length,
            seed=seed,
        )
        if data_mode == 'random':
            reference_source_label = 'Sinh ngẫu nhiên'
            reference_source_description = 'Hệ thống tự tạo chuỗi trang để demo các lỗi trang nhanh hơn.'

    comparison = build_memory_comparison(refs, frame_count=frame_count)
    algo_map = {row['name'].lower(): row for row in comparison['algorithms']}
    selected = algo_map.get(algorithm, comparison['algorithms'][1] if len(comparison['algorithms']) > 1 else comparison['algorithms'][0])

    selected_steps = selected['steps']
    steps_json = json.dumps([
        {
            'step': item['step'],
            'ref': item['ref'],
            'hit': item['hit'],
            'replaced': item['replaced'],
            'slot': item['slot'],
            'frames': item['frames'],
        } for item in selected_steps
    ])

    context = {
        'algorithm': selected['name'].lower(),
        'algorithm_name': selected['name'],
        'frame_count': frame_count,
        'data_mode': data_mode,
        'reference_source_label': reference_source_label,
        'reference_source_description': reference_source_description,
        'references': refs,
        'references_json': json.dumps(refs),
        'steps_json': steps_json,
        'selected': selected,
        'comparison': comparison['algorithms'],
        'comparison_json': json.dumps([
            {
                'name': row['name'],
                'faults': row['stats']['faults'],
                'hit_rate': row['stats']['hit_rate'],
                'fault_rate': row['stats']['fault_rate'],
            } for row in comparison['algorithms']
        ]),
        'best_faults': comparison['best_faults'],
        'best_hit': comparison['best_hit'],
        'manual_reference_raw': manual_reference_raw,
        'seed': seed,
        'ref_length': ref_length,
    }
    return render(request, 'app/memory_sim.html', context)


# ==============================================================
# MÔ PHỎNG — ĐỒNG BỘ HÓA
# ==============================================================
@login_required(login_url='app:login')
def synchronization_simulation(request):
    expire_overdue_sessions()

    buffer_size = max(1, int(request.GET.get('buffer_size') or (request.session.get('custom_sync', {}).get('buffer_size') if request.GET.get('pattern') == 'file' else 0) or 3))
    producer_count = max(1, int(request.GET.get('producer_count') or (request.session.get('custom_sync', {}).get('producer_count') if request.GET.get('pattern') == 'file' else 0) or 2))
    consumer_count = max(1, int(request.GET.get('consumer_count') or (request.session.get('custom_sync', {}).get('consumer_count') if request.GET.get('pattern') == 'file' else 0) or 2))
    step_count = max(6, int(request.GET.get('step_count', 12)))
    seed = request.GET.get('seed', '').strip()
    pattern = request.GET.get('pattern', 'auto').lower()
    total_ram = max(1, int(request.GET.get('total_ram', 512)))
    mem_producer = max(1, int(request.GET.get('mem_producer', 128)))
    mem_consumer = max(1, int(request.GET.get('mem_consumer', 96)))

    sessions = list(Session.objects.filter(status='dang_chay').select_related('machine', 'customer').order_by('start_time')[:20])
    queue_items = list(Queue.objects.filter(is_served=False).order_by('arrived_at')[:20])

    rng = random.Random(seed) if seed else random.Random()
    if pattern == 'manual':
        manual_sequence = request.GET.get('sequence', '')
        sequence = [token.strip().upper() for token in manual_sequence.replace(',', ' ').split() if token.strip()]
        sequence = [token[0] for token in sequence if token[0] in {'P', 'C'}]
    elif pattern == 'file':
        custom_sync = request.session.get('custom_sync', {})
        sequence = custom_sync.get('sequence', ['P', 'C'])
        step_count = len(sequence)
    else:
        sequence = []
        if pattern == 'auto':
            live_events = []
            live_events.extend((item.arrived_at, 'P') for item in queue_items)
            live_events.extend((sess.expected_end_time, 'C') for sess in sessions)
            sequence = [action for _, action in sorted(live_events, key=lambda row: row[0])[:step_count]]
        else:
            for i in range(step_count):
                if pattern == 'producer_first':
                    sequence.append('P' if i < step_count // 2 else 'C')
                elif pattern == 'consumer_first':
                    sequence.append('C' if i < step_count // 2 else 'P')
                else:
                    sequence.append(rng.choice(['P', 'C']))
            if sequence and 'P' not in sequence:
                sequence[0] = 'P'
            if sequence and 'C' not in sequence:
                sequence[-1] = 'C'

    result = simulate_bounded_buffer(
        buffer_size=buffer_size,
        sequence=sequence,
        producer_count=producer_count,
        consumer_count=consumer_count,
        total_ram=total_ram,
        mem_producer=mem_producer,
        mem_consumer=mem_consumer,
    )

    occupancy = [len(step['buffer']) for step in result['timeline']]
    ram_timeline = [step['ram_free'] for step in result['timeline']]
    sync_json = json.dumps([
        {
            'step': item['step'],
            'actor': item['actor'],
            'action': item['action'],
            'result': item['result'],
            'buffer': item['buffer'],
            'empty': item['empty'],
            'full': item['full'],
            'ram_free': item['ram_free'],
            'description': item['description'],
        } for item in result['timeline']
    ])

    context = {
        'buffer_size': buffer_size,
        'producer_count': producer_count,
        'consumer_count': consumer_count,
        'step_count': step_count,
        'seed': seed,
        'pattern': pattern,
        'total_ram': total_ram,
        'mem_producer': mem_producer,
        'mem_consumer': mem_consumer,
        'sequence': sequence,
        'result': result,
        'live_session_count': len(sessions),
        'live_queue_count': len(queue_items),
        'occupancy_json': json.dumps(occupancy),
        'ram_json': json.dumps(ram_timeline),
        'sync_json': sync_json,
    }
    return render(request, 'app/synchronization_sim.html', context)


# ==============================================================
# BÁO CÁO
# ==============================================================
@login_required(login_url='app:login')
def report(request):
    today = timezone.now().date()
    sessions_today = Session.objects.filter(start_time__date=today)

    revenue  = sessions_today.filter(paid=True).aggregate(t=Sum('total_cost'))['t'] or 0
    by_type  = sessions_today.filter(paid=True).values('machine__machine_type').annotate(
        count=Count('id'), total=Sum('total_cost')
    )

    active_sessions = Session.objects.filter(status='dang_chay').select_related('machine', 'customer')
    waiting_queue = Queue.objects.filter(is_served=False).order_by('arrived_at')

    scheduling_comparison = []
    reference_schedule = []
    if active_sessions or waiting_queue:
        reference_processes = sessions_to_processes(active_sessions[:5]) + queue_to_processes(waiting_queue[:5])
        normalize_arrival_times(reference_processes)
        if reference_processes:
            scheduling_comparison = [
                {'key': 'fcfs', 'name': 'FCFS', 'stats': calc_stats(fcfs(reference_processes), reference_processes)},
                {'key': 'sjf', 'name': 'SJF', 'stats': calc_stats(sjf(reference_processes), reference_processes)},
                {'key': 'rr', 'name': 'RR', 'stats': calc_stats(round_robin(reference_processes, 30), reference_processes)},
            ]
            reference_schedule = reference_processes[:6]

    memory_refs = build_memory_reference_stream(
        'live',
        sessions=active_sessions[:5],
        queue_items=waiting_queue[:5],
        length=12,
    )
    memory_comparison = build_memory_comparison(memory_refs, frame_count=3)

    implementation_summary = [
        {
            'title': 'Tiến trình',
            'status': 'Đã có',
            'detail': 'Mỗi phiên chơi được ánh xạ thành tiến trình với arrival time, burst time và resource vector.',
        },
        {
            'title': 'CPU Scheduling',
            'status': 'Đã có',
            'detail': 'Đã mô phỏng FCFS, SJF, Round Robin và có so sánh hiệu năng bằng WT, TAT, Response, CPU Utilization.',
        },
        {
            'title': 'Bộ nhớ',
            'status': 'Đã bổ sung',
            'detail': 'Có mô phỏng FIFO, LRU, OPT với biểu đồ lỗi trang và tỷ lệ hit/fault.',
        },
        {
            'title': 'Đồng bộ hóa',
            'status': 'Đã bổ sung',
            'detail': 'Có mô phỏng bounded buffer với producer-consumer, semaphore và vùng găng.',
        },
        {
            'title': 'Banker / Deadlock',
            'status': 'Đã có',
            'detail': 'Có kiểm tra an toàn cấp phát tài nguyên và mô phỏng xử lý deadlock.',
        },
    ]

    design_notes = [
        'Frontend Django template đóng vai trò giao diện mô phỏng và trình bày báo cáo.',
        'Dữ liệu thực của quán net được ánh xạ thành tiến trình, tài nguyên và hàng chờ.',
        'Các thuật toán được viết thuần Python để dễ demo, so sánh và giải thích trong bài.',
        'Giao diện dùng biểu đồ, bảng và animation để phù hợp yêu cầu trực quan.',
    ]

    research_highlights = [
        {
            'title': 'Chống deadlock bằng thứ tự tài nguyên',
            'detail': 'Nếu mỗi máy / tài nguyên được đánh số và tiến trình luôn xin theo thứ tự tăng dần, vòng chờ sẽ bị phá vỡ.',
            'source': 'CS 537 deadlock notes',
            'url': 'https://pages.cs.wisc.edu/~solomon/cs537/html/deadlock.html',
        },
        {
            'title': 'Banker cho trạng thái an toàn',
            'detail': 'Khi có yêu cầu mới, hệ thống mô phỏng cấp phát trước rồi chỉ chấp nhận nếu vẫn còn safe sequence.',
            'source': 'RIT Banker assignment',
            'url': 'https://www.se.rit.edu/~se441/Assignments/Bankers_Algorithm.html',
        },
        {
            'title': 'CPU scheduling có chống starvation',
            'detail': 'MLFQ kết hợp priority boost định kỳ giúp jobs interactive phản hồi nhanh mà vẫn không bỏ rơi job dài.',
            'source': 'OSTEP MLFQ + CS 537',
            'url': 'https://research.cs.wisc.edu/wind/OSTEP/cpu-sched-mlfq.pdf',
        },
        {
            'title': 'Memory tránh thrashing',
            'detail': 'Working set và page-fault frequency là hướng nâng cấp thực tế để cấp phát frame theo mức áp lực bộ nhớ.',
            'source': 'Stanford CS140',
            'url': 'https://web.stanford.edu/~ouster/cgi-bin/cs140-winter12/lecture.php?topic=thrashing',
        },
        {
            'title': 'Admission control cho 10 máy / 11 khách',
            'detail': 'Thay vì nhận tràn, hệ thống giữ khách thứ 11 ở hàng chờ và chỉ cấp phát khi còn tài nguyên an toàn.',
            'source': 'USENIX Redline',
            'url': 'https://www.usenix.org/legacy/events/osdi08/tech/full_papers/yang/yang_html/index.html',
        },
    ]

    context = {
        'today':          today,
        'sessions_today': sessions_today,
        'revenue':        revenue,
        'by_type':        by_type,
        'total_count':    sessions_today.count(),
        'active_sessions': active_sessions,
        'waiting_queue': waiting_queue,
        'scheduling_comparison': scheduling_comparison,
        'memory_comparison': memory_comparison['algorithms'],
        'memory_best_faults': memory_comparison['best_faults'],
        'memory_best_hit': memory_comparison['best_hit'],
        'implementation_summary': implementation_summary,
        'design_notes': design_notes,
        'reference_schedule': reference_schedule,
        'research_highlights': research_highlights,
    }
    return render(request, 'app/report.html', context)


# ==============================================================
# OS LAB+ - PHÂN TÍCH NÂNG CAO
# ==============================================================
@login_required(login_url='app:login')
def os_lab(request):
    expire_overdue_sessions()

    active_sessions = list(
        Session.objects.filter(status='dang_chay')
        .select_related('machine', 'customer')
        .order_by('start_time')[:8]
    )
    waiting_queue = list(
        Queue.objects.filter(is_served=False)
        .order_by('arrived_at')[:8]
    )
    machines = Machine.objects.all()

    process_pool = sessions_to_processes(active_sessions[:5]) + queue_to_processes(waiting_queue[:5])
    normalize_arrival_times(process_pool)
    scheduling_recommendation = {
        'best': 'FCFS',
        'why': 'Dữ liệu hiện tại ít và ổn định, FCFS dễ trình bày, trực quan, phù hợp demo live.',
        'comparison': [],
    }
    if process_pool:
        fcfs_tl = fcfs(process_pool)
        sjf_tl = sjf(process_pool)
        rr_tl = round_robin(process_pool, 30)
        scheduling_recommendation['comparison'] = [
            {'key': 'fcfs', 'name': 'FCFS', 'stats': calc_stats(fcfs_tl, process_pool)},
            {'key': 'sjf', 'name': 'SJF', 'stats': calc_stats(sjf_tl, process_pool)},
            {'key': 'rr', 'name': 'Round Robin', 'stats': calc_stats(rr_tl, process_pool)},
        ]
        scheduling_recommendation['best'] = min(
            scheduling_recommendation['comparison'],
            key=lambda row: row['stats']['avg_wait']
        )['name']
        scheduling_recommendation['why'] = (
            f"{scheduling_recommendation['best']} có thời gian chờ trung bình thấp nhất trong dữ liệu hiện tại, "
            "nên rất hợp để đưa vào phần so sánh trong báo cáo."
        )

    memory_refs = build_memory_reference_stream(
        'live',
        sessions=active_sessions[:5],
        queue_items=waiting_queue[:5],
        length=12,
    )
    memory_comparison = build_memory_comparison(memory_refs, frame_count=3)

    sync_events = []
    sync_events.extend((item.arrived_at, 'P') for item in waiting_queue[:12])
    sync_events.extend((sess.expected_end_time, 'C') for sess in active_sessions[:12])
    sync_sequence = [action for _, action in sorted(sync_events, key=lambda row: row[0])[:12]]
    sync_result = simulate_bounded_buffer(
        buffer_size=3,
        sequence=sync_sequence,
        producer_count=max(1, len(waiting_queue[:12])),
        consumer_count=max(1, len(active_sessions[:12])),
    )

    banker_data = build_banker_data(active_sessions)
    banker_safe = {
        'safe': True,
        'safe_sequence': [],
        'deadlock': False,
    }
    if banker_data['n'] > 0:
        banker_safe_result = banker_safety(
            banker_data['n'],
            banker_data['available'],
            banker_data['allocation'],
            banker_data['need'],
            session_names=[f"S{s.id} - {s.get_customer_name()}" for s in active_sessions],
        )
        banker_safe = {
            'safe': banker_safe_result.get('safe', True),
            'safe_sequence': banker_safe_result.get('safe_sequence', []),
            'deadlock': banker_safe_result.get('deadlock_info', {}).get('is_deadlock', False),
        }

    module_cards = [
        {
            'title': 'CPU Scheduling',
            'accent': 'CPU',
            'value': f"{len(process_pool)} tiến trình",
            'detail': 'So sánh FCFS, SJF và RR trên cùng một tập dữ liệu live.',
        },
        {
            'title': 'Memory',
            'accent': 'MEM',
            'value': f"{len(memory_refs)} lần truy cập",
            'detail': 'FIFO, LRU, OPT giúp có biểu đồ fault/hit rõ ràng cho báo cáo.',
        },
        {
            'title': 'Synchronization',
            'accent': 'SYNC',
            'value': f"{len(sync_result['timeline'])} bước",
            'detail': 'Producer-Consumer tạo animation vùng đệm đủ bắt mắt để thuyết trình.',
        },
        {
            'title': 'Banker',
            'accent': 'SAFE',
            'value': 'Kiểm tra an toàn',
            'detail': 'Có thể giải thích deadlock và state an toàn ngay trên dữ liệu thật.',
        },
    ]

    enhancement_points = [
        'Đã gắn dữ liệu thật của quán net vào mô phỏng thay vì chỉ dùng ví dụ tĩnh.',
        'Có so sánh nhiều thuật toán và chỉ ra lựa chọn tốt nhất theo tiêu chí hiện tại.',
        'Có trang phân tích riêng để trình bày phần mở rộng ngoài rubric.',
        'Có hệ thống màu riêng cho từng module để người xem đọc rất nhanh.',
        'Có thể dùng thẳng trong phần demo: CPU, Memory, Sync và Banker đều có kết quả trực quan.',
    ]

    context = {
        'active_sessions': active_sessions,
        'waiting_queue': waiting_queue,
        'machines': machines,
        'module_cards': module_cards,
        'enhancement_points': enhancement_points,
        'scheduling_recommendation': scheduling_recommendation,
        'memory_comparison': memory_comparison['algorithms'],
        'memory_best_faults': memory_comparison['best_faults'],
        'memory_best_hit': memory_comparison['best_hit'],
        'sync_result': sync_result,
        'sync_sequence': sync_sequence,
        'banker_safe': banker_safe,
    }
    return render(request, 'app/os_lab.html', context)


# ==============================================================
# XÁC THỰC
# ==============================================================
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            messages.success(request, f'Xin chào {username}! Chào mừng trở lại quán.')
            return redirect('app:dashboard')
        else:
            messages.error(request, 'Tên đăng nhập hoặc mật khẩu không đúng.')
    return render(request, 'app/login.html')


def phone_login_request(request):
    if request.method == 'POST':
        form = PhoneRequestForm(request.POST)
        if form.is_valid():
            phone = form.cleaned_data['phone']
            code = f"{random.randint(100000,999999)}"
            OTPCode.objects.create(phone=phone, code=code)
            # Dev-mode: show code in message (in production send SMS)
            messages.success(request, f'OTP cho {phone}: {code} (dev-mode)')
            return redirect('app:phone_verify')
    else:
        form = PhoneRequestForm()
    return render(request, 'app/auth/phone_request.html', {'form': form})


def phone_login_verify(request):
    if request.method == 'POST':
        form = OTPVerifyForm(request.POST)
        if form.is_valid():
            phone = form.cleaned_data['phone']
            code = form.cleaned_data['code']
            otp = OTPCode.objects.filter(phone=phone, code=code, used=False).order_by('-created').first()
            if otp and otp.is_valid():
                otp.used = True
                otp.save()
                User = get_user_model()
                username = f'phone_{phone}'
                user, created = User.objects.get_or_create(username=username, defaults={'is_active': True})
                # Create or link Customer
                customer, _ = Customer.objects.get_or_create(phone=phone, defaults={'name': phone})
                # Log the user in
                user.backend = 'django.contrib.auth.backends.ModelBackend'
                login(request, user)
                messages.success(request, 'Đăng nhập thành công')
                return redirect('app:dashboard')
            messages.error(request, 'Mã OTP không hợp lệ hoặc đã hết hạn')
    else:
        form = OTPVerifyForm()
    return render(request, 'app/auth/phone_verify.html', {'form': form})


def qr_token_create(request):
    # Create a short token and show URL (dev-mode)
    token = ''.join(random.choices(string.ascii_letters + string.digits, k=32))
    lt = LoginToken.objects.create(token=token)
    url = request.build_absolute_uri(f'/auth/qr-login/{token}/')
    return render(request, 'app/auth/qr_create.html', {'token': token, 'url': url})


def qr_login(request, token):
    lt = get_object_or_404(LoginToken, token=token)
    if not lt.is_valid():
        messages.error(request, 'Token không hợp lệ hoặc đã hết hạn')
        return redirect('app:dashboard')
    # For dev: if phone present, log that user in; otherwise create anonymous user
    if lt.phone:
        User = get_user_model()
        username = f'phone_{lt.phone}'
        user, _ = User.objects.get_or_create(username=username)
        user.backend = 'django.contrib.auth.backends.ModelBackend'
        login(request, user)
        lt.used = True
        lt.save()
        messages.success(request, 'Đăng nhập bằng QR thành công')
        return redirect('app:dashboard')
    messages.info(request, 'Token tạo thành công. Bạn có thể liên kết số điện thoại để đăng nhập.')
    return render(request, 'app/auth/qr_login.html', {'token': token})


def logout_view(request):
    logout(request)
    messages.info(request, 'Đã đăng xuất thành công.')
    return redirect('app:login')


def register_view(request):
    form = UserCreationForm()
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Đăng ký tài khoản nhân viên thành công!')
            return redirect('app:dashboard')
        else:
            messages.error(request, 'Thông tin đăng ký không hợp lệ.')
    return render(request, 'app/register.html', {'form': form})


# ==============================================================
# STUB VIEWS — E-COMMERCE ROUTES (for template compatibility)
# ==============================================================
@login_required(login_url='app:login')
def product_list(request):
    """Placeholder for product list view"""
    return render(request, 'app/product/list.html', {'products': []})

@login_required(login_url='app:login')
def product_detail(request, product_id, slug=''):
    """Placeholder for product detail view"""
    return render(request, 'app/product/detail.html', {'product': None})

@login_required(login_url='app:login')
def add_to_cart(request, product_id):
    """Placeholder for add to cart"""
    return redirect('app:view_cart')

@login_required(login_url='app:login')
def view_cart(request):
    """Placeholder for cart view"""
    return render(request, 'app/product/cart.html', {'items': []})

@login_required(login_url='app:login')
def remove_item(request, item_id):
    """Placeholder for remove cart item"""
    return redirect('app:view_cart')

def about(request):
    """Placeholder for about page"""
    return render(request, 'app/about.html', {})

def contact(request):
    """Placeholder for contact page"""
    return render(request, 'app/contact.html', {})

def search(request):
    """Placeholder for search"""
    return render(request, 'app/product/search.html', {'results': []})
